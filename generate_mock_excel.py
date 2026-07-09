#!/usr/bin/env python3
"""
Script to generate mock Excel files for testing Async Reconciliation Architecture V2.
Uses openpyxl to create realistic data structures matching the original scripts' expectations.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timedelta
import random


def create_padroeira_vendas():
    """Create Movto_cx2.xlsx for Padroeira_vendas with data until July 6, 2026."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimento"

    # Header
    headers = [
        "Data", "Dia", "Dinheiro", "Crédito", "Débito", "Total",
        "Clientes", "Média", "Buf", "Sob", "Saldo", "Observações"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Data from June 1 to July 6, 2026
    start_date = datetime(2026, 6, 1)
    end_date = datetime(2026, 7, 6)
    current_date = start_date

    while current_date <= end_date:
        day = current_date.day
        weekday = current_date.strftime("%A")

        # Generate realistic values
        dinheiro = random.randint(800, 1500)
        credito = random.randint(2000, 4000)
        debito = random.randint(1500, 3500)
        total = dinheiro + credito + debito
        clientes = random.randint(30, 60)
        media = round(total / clientes, 2) if clientes > 0 else 0
        buf = random.randint(20, 100)
        sob = random.randint(10, 50)
        saldo = total - buf + sob

        row = [
            current_date.strftime("%d/%m/%Y"),
            weekday,
            dinheiro,
            credito,
            debito,
            total,
            clientes,
            media,
            buf,
            sob,
            saldo,
            ""  # Observações
        ]
        ws.append(row)
        current_date += timedelta(days=1)

    # Add SANGRIA at row 42 (after June data)
    ws.insert_rows(42)
    ws.merge_cells(start_row=42, start_column=1, end_row=42, end_column=12)
    ws.cell(row=42, column=1, value="SANGRIA").font = Font(bold=True, size=14)
    ws.cell(row=42, column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Save
    wb.save("/home/teco/work_out/mock_box/Padroeira_vendas/Movto_cx2.xlsx")


def create_restaurante_movto_diario():
    """Create Movto_diario.2607.xlsx for Restaurante stopped at June 28, 2026."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimento Diário"

    # Header
    headers = [
        "Data", "Dia", "Receita", "Despesa", "Saldo", "Clientes",
        "Média", "Observações"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Data from June 1 to June 28, 2026
    start_date = datetime(2026, 6, 1)
    end_date = datetime(2026, 6, 28)
    current_date = start_date

    while current_date <= end_date:
        day = current_date.day
        weekday = current_date.strftime("%A")

        # Generate realistic values
        receita = random.randint(3000, 6000)
        despesa = random.randint(1000, 2500)
        saldo = receita - despesa
        clientes = random.randint(50, 120)
        media = round(receita / clientes, 2) if clientes > 0 else 0

        row = [
            current_date.strftime("%d/%m/%Y"),
            weekday,
            receita,
            despesa,
            saldo,
            clientes,
            media,
            ""  # Observações
        ]
        ws.append(row)
        current_date += timedelta(days=1)

    # Save
    wb.save("/home/teco/work_out/mock_box/Restaurante/A2026/Movto_diario.2607.xlsx")


def create_restaurante_pad():
    """Create Pad2607.xlsx for Restaurante with July rows empty or stopped at day 20."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PAD"

    # Header
    headers = [
        "Data", "Dia", "Receita", "Despesa", "Saldo", "Clientes",
        "Média", "Buf", "Sob", "Observações"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Data from June 1 to July 20, 2026 (July rows empty)
    start_date = datetime(2026, 6, 1)
    end_date = datetime(2026, 7, 20)
    current_date = start_date

    while current_date <= end_date:
        day = current_date.day
        weekday = current_date.strftime("%A")

        # June data filled, July data empty
        if current_date.month == 6:
            receita = random.randint(3000, 6000)
            despesa = random.randint(1000, 2500)
            saldo = receita - despesa
            clientes = random.randint(50, 120)
            media = round(receita / clientes, 2) if clientes > 0 else 0
            buf = random.randint(50, 200)
            sob = random.randint(20, 100)
        else:
            receita = despesa = saldo = clientes = media = buf = sob = None

        row = [
            current_date.strftime("%d/%m/%Y"),
            weekday,
            receita,
            despesa,
            saldo,
            clientes,
            media,
            buf,
            sob,
            ""  # Observações
        ]
        ws.append(row)
        current_date += timedelta(days=1)

    # Save
    wb.save("/home/teco/work_out/mock_box/Restaurante/A2026/Pad2607.xlsx")


if __name__ == "__main__":
    create_padroeira_vendas()
    create_restaurante_movto_diario()
    create_restaurante_pad()
    print("Mock Excel files generated successfully.")