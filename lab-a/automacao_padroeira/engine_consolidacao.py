#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Engine de Consolidação - Mestre Padroeira
Motor Unificado: Expansão de Calendário + Espelhamento Vertical
Córtex de Desenvolvimento: work_out/lab
"""

import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

RAIZ_PADROEIRA  = "/home/teco/Nuvens/Box/Padroeira"
PASTA_VENDAS    = os.path.join(RAIZ_PADROEIRA, "Padroeira vendas")
PASTA_RESTAURANTE_ANO = os.path.join(RAIZ_PADROEIRA, "Restaurante", "A2026")

AAMM = "2606"  # Junho

def normalizar_data(valor):
    """Retorna um objeto date para casamento seguro de colunas"""
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, str):
        try:
            return datetime.strptime(valor.strip().split()[0], "%d/%m/%Y").date()
        except:
            try:
                return datetime.strptime(valor.strip().split()[0], "%Y-%m-%d").date()
            except:
                return None
    return None

def executar_motor_unificado():
    caminho_caixa2 = os.path.join(PASTA_VENDAS, "Movto_cx2.xlsx")
    caminho_mensal = os.path.join(PASTA_RESTAURANTE_ANO, f"Movto_diario.{AAMM}.xlsx")
    
    if not os.path.exists(caminho_caixa2) or not os.path.exists(caminho_mensal):
        print("[-] Erro Crítico: Arquivos base não encontrados.")
        return

    print("[*] Carregando matrizes na memória (Isso pode levar alguns segundos)...")
    
    # Caixa 2 (Leitura dos dados da Sandra)
    wb_cx = load_workbook(caminho_caixa2, data_only=True)
    ws_cx = wb_cx.active

    # Diário Mensal (Escrita - preserva fórmulas)
    wb_me = load_workbook(caminho_mensal, data_only=False)
    ws_me = wb_me.active

    # Diário Mensal (Leitura - checa resultados das fórmulas existentes)
    wb_me_lei = load_workbook(caminho_mensal, data_only=True)
    ws_me_lei = wb_me_lei.active

    # ==============================================================================
    # ETAPA 1: MAPEAMENTO E EXPANSÃO HORIZONTAL
    # ==============================================================================
    datas_no_mensal = {}
    colunas_pendentes_de_carga = [] # Guarda os índices das colunas que precisarão de injeção
    proxima_coluna_livre = 2
    
    # 1.1 Varre o Diário atual para entender o que já existe
    for col in range(2, ws_me_lei.max_column + 1):
        v_data = ws_me_lei.cell(row=1, column=col).value
        v_total = ws_me_lei.cell(row=24, column=col).value # Linha de controle
        
        if v_data:
            dt_norm = normalizar_data(v_data)
            if dt_norm:
                datas_no_mensal[dt_norm] = col
                # Se o dia existe, mas a linha 24 está vazia/zerada, agenda para receber carga
                if v_total is None or float(v_total) == 0:
                    colunas_pendentes_de_carga.append(col)
                    
        # Identifica a borda da matriz para criar novas datas
        if ws_me_lei.cell(row=1, column=col).value is None and proxima_coluna_livre == 2:
            proxima_coluna_livre = col

    if proxima_coluna_livre == 2:
        proxima_coluna_livre = ws_me_lei.max_column + 1

    # 1.2 Lê o Caixa 2 e separa as datas do mês ativo
    mes_alvo = int(AAMM[2:])
    ano_alvo = 2000 + int(AAMM[:2])
    mapa_cx_datas = {}
    datas_faltantes = []

    for col in range(2, ws_cx.max_column + 1):
        v_data_cx = ws_cx.cell(row=1, column=col).value
        if v_data_cx:
            dt_cx_norm = normalizar_data(v_data_cx)
            if dt_cx_norm and dt_cx_norm.month == mes_alvo and dt_cx_norm.year == ano_alvo:
                mapa_cx_datas[dt_cx_norm] = col
                if dt_cx_norm not in datas_no_mensal and dt_cx_norm not in datas_faltantes:
                    datas_faltantes.append(dt_cx_norm)

    datas_faltantes.sort()

    # 1.3 Injeta as colunas novas na Linha 1 do Diário (se houver)
    if datas_faltantes:
        print(f"[*] Expandindo o calendário: {len(datas_faltantes)} novos dias detectados.")
        col_atual = proxima_coluna_livre
        for nova_data in datas_faltantes:
            dt_objeto = datetime(nova_data.year, nova_data.month, nova_data.day)
            ws_me.cell(row=1, column=col_atual, value=dt_objeto)
            ws_me.cell(row=1, column=col_atual).number_format = 'd-mmm-yy'
            
            # Registra no mapa e já agenda a nova coluna para receber carga
            datas_no_mensal[nova_data] = col_atual
            colunas_pendentes_de_carga.append(col_atual)
            col_atual += 1
    else:
        print("[+] A matriz de calendário já está sincronizada.")

    # ==============================================================================
    # ETAPA 2: ESPELHAMENTO VERTICAL (INJEÇÃO DE DADOS E FÓRMULAS)
    # ==============================================================================
    if not colunas_pendentes_de_carga:
        print("[+] Paridade de faturamento total! Nenhuma coluna precisa de dados.")
    else:
        print(f"\n[*] Iniciando injeção em {len(colunas_pendentes_de_carga)} colunas pendentes...")
        mudancas = 0
        
        # Filtra o mapa para garantir que temos o dado no Caixa 2
        dias_para_processar = []
        for dt, col_me in datas_no_mensal.items():
            if col_me in colunas_pendentes_de_carga and dt in mapa_cx_datas:
                dias_para_processar.append({
                    "data": dt,
                    "col_caixa": mapa_cx_datas[dt],
                    "col_mensal": col_me
                })

        # Executa o paralelismo vertical
        for dia in dias_para_processar:
            letra_col_me = get_column_letter(dia['col_mensal'])
            print(f"    -> Transportando dados para {dia['data'].strftime('%d/%m/%Y')} (Coluna {letra_col_me})")
            
            for linha in range(6, ws_cx.max_row + 1):
                # Preserva fórmulas vitais
                if linha == 14:
                    ws_me.cell(row=linha, column=dia['col_mensal'], value=f"=SUM({letra_col_me}10:{letra_col_me}13)")
                    mudancas += 1
                    continue
                if linha == 40:
                    ws_me.cell(row=linha, column=dia['col_mensal'], value=f"={letra_col_me}37-{letra_col_me}38")
                    mudancas += 1
                    continue
                    
                # Injeta valor bruto
                valor_real_cx = ws_cx.cell(row=linha, column=dia['col_caixa']).value
                ws_me.cell(row=linha, column=dia['col_mensal'], value=valor_real_cx)
                mudancas += 1

        print(f"[+] Espelhamento concluído! {mudancas} células tratadas.")

    # ==============================================================================
    # ETAPA 3: SALVAMENTO ESTRATÉGICO
    # ==============================================================================
    try:
        wb_me.save(caminho_mensal)
        print(f"\n[SUCESSO] Arquivo Movto_diario.{AAMM}.xlsx finalizado e salvo na nuvem!")
    except Exception as e:
        print(f"[-] Erro Crítico ao salvar o arquivo Diário: {e}")

if __name__ == "__main__":
    print("[*] ==================================================================")
    print("[*] CÓRTEX PADROEIRA - MOTOR UNIFICADO (DIÁRIO MENSAL)")
    print("[*] ==================================================================")
    executar_motor_unificado()
