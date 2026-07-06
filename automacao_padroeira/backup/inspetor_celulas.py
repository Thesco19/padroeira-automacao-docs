#!/usr/bin/env python3
import os
from openpyxl import load_workbook

RAIZ = "/home/teco/Nuvens/Box/Padroeira"
cx2 = os.path.join(RAIZ, "Padroeira vendas", "Movto_cx2.xlsx")
mes = os.path.join(RAIZ, "Restaurante", "A2026", "Movto_diario.2606.xlsx")

def inspecionar(caminho, nome):
    if not os.path.exists(caminho):
        print(f"[-] {nome} não encontrado em {caminho}")
        return
    wb = load_workbook(caminho, data_only=True)
    ws = wb.active
    print(f"\n[*] INSPEÇÃO FÍSICA: {nome} ({ws.max_column} colunas detectadas)")
    
    # Mostra o que está gravado nas 15 primeiras colunas (Linha 1 e Linha 24)
    for col in range(1, min(15, ws.max_column + 1)):
        data_celula = ws.cell(row=1, column=col).value
        tipo_data = type(data_celula).__name__
        total_celula = ws.cell(row=24, column=col).value
        print(f"    Col {col} -> Linha 1 (Data): {data_celula} [{tipo_data}] | Linha 24 (Total): {total_celula} ({type(total_celula).__name__})")

if __name__ == "__main__":
    inspecionar(cx2, "CAIXA 2")
    inspecionar(mes, "MENSAL JUNHO")
