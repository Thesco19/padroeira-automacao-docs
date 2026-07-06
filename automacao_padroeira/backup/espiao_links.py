#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
from openpyxl import load_workbook

# Apontando para o seu Balancete real do espelho do Box
CAMINHO_BALANCETE = "/home/teco/Nuvens/Box/Padroeira/Restaurante/A2026/Movto_diario.2605"

# Ajuste caso tenha extensão
if not os.path.exists(CAMINHO_BALANCETE) and os.path.exists(CAMINHO_BALANCETE + ".xlsx"):
    CAMINHO_BALANCETE += ".xlsx"

def espiar_links():
    if not os.path.exists(CAMINHO_BALANCETE):
        print(f"[-] Arquivo não encontrado: {CAMINHO_BALANCETE}")
        return

    print(f"[*] Abrindo {CAMINHO_BALANCETE} para analisar fórmulas...")
    wb = load_workbook(CAMINHO_BALANCETE, data_only=False) # data_only=False traz a fórmula pura
    
    print("[*] Buscando links externos nas células...")
    links_encontrados = set()
    
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "[" in cell.value:
                    # Captura a menção a outros arquivos (ex: [movto_cx1_1508.XLSX])
                    links_encontrados.add(cell.value)
                    
    if links_encontrados:
        print("\n[+] Links externos detectados dentro do balancete:")
        for link in list(links_encontrados)[:5]: # Mostra os 5 primeiros exemplos
            print(f"    -> {link}")
    else:
        print("[-] Nenhum link externo explícito foi encontrado em formato de texto.")

if __name__ == "__main__":
    espiar_links()
