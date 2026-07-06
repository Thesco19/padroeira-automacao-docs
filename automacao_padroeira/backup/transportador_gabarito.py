#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Engine de Consolidação - Módulo 2: Transporte por Espelhamento e Fórmulas Dinâmicas
Córtex de Desenvolvimento: work_out/lab
"""

import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

RAIZ_PADROEIRA  = "/home/teco/Nuvens/Box/Padroeira"
PASTA_VENDAS    = os.path.join(RAIZ_PADROEIRA, "Padroeira vendas")
PASTA_RESTAURANTE_ANO = os.path.join(RAIZ_PADROEIRA, "Restaurante", "A2026")

AAMM = "2606"  # Junho

def normalizar_data(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, str):
        try:
            return datetime.strptime(valor.strip().split()[0], "%d/%m/%Y").date()
        except:
            try:
                return datetime.strptime(valor.strip().split()[0], "%Y-%m-%d").date()
            except:
                return None
    return None

def rodar_transporte_espelhado():
    cx2_path = os.path.join(PASTA_VENDAS, "Movto_cx2.xlsx")
    mensal_path = os.path.join(PASTA_RESTAURANTE_ANO, f"Movto_diario.{AAMM}.xlsx")
    
    if not os.path.exists(cx2_path) or not os.path.exists(mensal_path):
        print("[-] Erro: Arquivos base não localizados no Box.")
        return

    print("[*] Iniciando Fase 2: Mapeamento horizontal por pareamento estrito...")
    
    # Abre o Mensal em modo LEITURA para capturar as datas e checar faturamento
    wb_me_lei = load_workbook(mensal_path, data_only=True)
    ws_me_lei = wb_me_lei.active

    # Abre o Caixa 2 em modo LEITURA para sugar os valores da Sandra
    wb_cx = load_workbook(cx2_path, data_only=True)
    ws_cx = wb_cx.active

    # 1. Identifica quais colunas do Diário Mensal estão zeradas/vazias na Linha 24
    colunas_alvo_mensal = {}
    for col in range(2, ws_me_lei.max_column + 1):
        v_data = ws_me_lei.cell(row=1, column=col).value
        v_total = ws_me_lei.cell(row=24, column=col).value
        
        if v_data:
            dt_norm = normalizar_data(v_data)
            # Modificação tática: se a linha 24 for 0 ou None, a coluna precisa receber dados
            if dt_norm and (v_total is None or float(v_total) == 0):
                colunas_alvo_mensal[dt_norm] = col

    if not colunas_alvo_mensal:
        print("[+] Tudo atualizado! Nenhuma coluna vazia detectada no Diário de Junho.")
        return

    print(f"[*] Dias identificados para recebimento de carga: {[d.strftime('%d/%m/%Y') for d in colunas_alvo_mensal.keys()]}")

    # 2. Encontra as colunas correspondentes dentro do Caixa 2
    mapa_pontes = []
    for col_cx in range(2, ws_cx.max_column + 1):
        v_data_cx = ws_cx.cell(row=1, column=col_cx).value
        if v_data_cx:
            dt_cx_norm = normalizar_data(v_data_cx)
            if dt_cx_norm in colunas_alvo_mensal:
                mapa_pontes.append({
                    "data": dt_cx_norm,
                    "col_caixa": col_cx,
                    "col_mensal": colunas_alvo_mensal[dt_cx_norm]
                })

    if not mapa_pontes:
        print("[-] Nenhuma das colunas vazias do Diário possui correspondente no Caixa 2 por enquanto.")
        return

    # 3. Abre o arquivo Mensal em modo ESCRITA (data_only=False) para salvar as fórmulas
    wb_me_escrita = load_workbook(mensal_path, data_only=False)
    ws_me_esc = wb_me_escrita.active

    print(f"\n[*] [TRANSPORTE ESPELHADO] Processando {len(mapa_pontes)} dias com injeção de fórmulas...")
    mudancas = 0

    for dia in mapa_pontes:
        letra_col_me = get_column_letter(dia['col_mensal'])
        print(f"    -> Sincronizando Dia {dia['data'].strftime('%d/%m/%Y')} (Letra Coluna Diário: {letra_col_me})")
        
        # Varre verticalmente a partir da Linha 6 até o limite físico da planilha
        for linha in range(6, ws_cx.max_row + 1):
            
            # TRATAMENTO ESPECIAL 1: Linha 14 (Subtotal da 10 à 13) -> Grava a fórmula!
            if linha == 14:
                formula_subtotal = f"=SUM({letra_col_me}10:{letra_col_me}13)"
                ws_me_esc.cell(row=linha, column=dia['col_mensal'], value=formula_subtotal)
                mudancas += 1
                continue
                
            # TRATAMENTO ESPECIAL 2: Linha 40 (Diferença L37 - L38) -> Grava a fórmula!
            if linha == 40:
                formula_diferenca = f"={letra_col_me}37-{letra_col_me}38"
                ws_me_esc.cell(row=linha, column=dia['col_mensal'], value=formula_diferenca)
                mudancas += 1
                continue
                
            # FLUXO PADRÃO: Linhas comuns espelhadas de 6 até o final
            valor_real_cx = ws_cx.cell(row=linha, column=dia['col_caixa']).value
            
            # Injeta o valor do caixa diretamente na linha idêntica do diário
            ws_me_esc.cell(row=linha, column=dia['col_mensal'], value=valor_real_cx)
            mudancas += 1

    # 4. Gravação em disco do Box
    if mudancas > 0:
        try:
            wb_me_escrita.save(mensal_path)
            print(f"\n[+] Sucesso Absoluto! {mudancas} células tratadas. Linhas 14 e 40 blindadas com fórmulas nativas.")
        except Exception as e:
            print(f"[-] Erro ao salvar o Diário Mensal: {e}")
    else:
        print("[-] Nenhuma modificação aplicada nas matrizes.")

if __name__ == "__main__":
    print("[*] ==================================================================")
    print("[*] PHASE 1.2: PIPELINE DE TRANSPORTE VIA PARALELISMO VERTICAL")
    print("[*] ==================================================================")
    rodar_transporte_espelhado()
