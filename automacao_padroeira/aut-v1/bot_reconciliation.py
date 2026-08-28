#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Bot de Reconciliação Padroeira - Ponto Único (bot escuta -> orquestra).

Este é o ELO que faltava no ecossistema: um bot Telegram que ESCUTA o comando
e ORQUESTRA todo o pipeline (Cortex -> Engine -> Balancete) para um ou mais
períodos AAMM, usando a extração SESSÃO ÚNICA do Saurus (reaproveitada do
`extrator_saurus_sessao`, a versão comprovada que produziu 264/0 relatórios).

Comandos:
    /fechar              -> PRIMEIRO comando do operador: puxa o FATURAMENTO DO DIA
                             (linha 37 "Real" do Diário), lê para a CONFERÊNCIA DE CAIXA
                             e SALVA NO HISTÓRICO para uso futuro. Apenas leitura/gravação
                             de histórico — não roda o pipeline nem toca o balancete.
    /finalizar [MMAA]    -> CONCLUI o preenchimento e o transporte de dados
                             (Cortex -> Engine -> Balancete Pad). Sem MMAA, assume o dia
                             de hoje (AAMM corrente). Se MMAA informado, faz a varredura
                             completa do período. /reconciliar é mantido como alias.
    /amostra [N] [MMAA]  -> roda apenas N datas pendentes (default 3) — útil p/ teste e2e.

Logs em tempo real: reconciliation.log é zerado a cada start (mode "w") e
espelhado no stdout. Marcadores exatos exigidos pelo teste de produção:
    [TELEGRAM] Comando recebido do usuário.
    [PLAYWRIGHT] Baixando fechamento para a data DD/MM/AAAA...
    [ENGINE] Injetando Kg Equivalente e Sangria (Linha 42) em Movto_diario.AAMM.xlsx...
    [TELEGRAM] Mensagem de resumo enviada ao usuário.
"""

import asyncio
import importlib.util
import json
import logging
import os
import re
import sys
from datetime import datetime

from openpyxl import load_workbook

# ----------------------------------------------------------------------
# Logging: FileHandler (zera o arquivo a cada start) + StreamHandler, DEBUG.
# ----------------------------------------------------------------------
WORK = os.path.dirname(os.path.abspath(__file__))            # aut-v1 (raiz do projeto; xlsx de teste ficam aqui)
PARENT = os.path.dirname(WORK)                                # automation_padroeira (pai)

# Histórico de faturamento do dia (lido pelo /fechar, persistido para uso futuro).
HIST_DIR = os.path.join(WORK, "historico_faturamento")
HIST_FILE = os.path.join(HIST_DIR, "faturamento_diario.json")

os.makedirs(os.path.join(WORK, "logs"), exist_ok=True)
os.makedirs(HIST_DIR, exist_ok=True)
LOGFILE = os.path.join(WORK, "logs", "reconciliation.log")

logger = logging.getLogger("BotReconciliation")
logger.setLevel(logging.DEBUG)

_fh = logging.FileHandler(LOGFILE, mode="w", encoding="utf-8")
_fh.setLevel(logging.DEBUG)
_sh = logging.StreamHandler(sys.stdout)
_sh.setLevel(logging.DEBUG)

_fmt = logging.Formatter(
    "%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
_fh.setFormatter(_fmt)
_sh.setFormatter(_fmt)
logger.addHandler(_fh)
logger.addHandler(_sh)


# ----------------------------------------------------------------------
# Carga de módulos via importlib (mesmo padrão de pre_producao_2608.py).
# ----------------------------------------------------------------------
def _carregar(modulo: str, caminho: str):
    spec = importlib.util.spec_from_file_location(modulo, caminho)
    m = importlib.util.module_from_spec(spec)
    sys.modules[modulo] = m
    spec.loader.exec_module(m)
    return m


sys.path.insert(0, PARENT)

cortex_mod = _carregar("cortex_padroeira_async", os.path.join(WORK, "cortex_padroeira_async.py"))
engine_mod = _carregar("engine_consolidacao_async", os.path.join(WORK, "engine_consolidacao_async.py"))
bal_mod = _carregar("motor_balancete_async", os.path.join(WORK, "motor_balancete_async.py"))
async_recon_mod = _carregar("async_reconciliation_v2", os.path.join(WORK, "async_reconciliation_v2.py"))

# Força o BASE_DIR do engine para a pasta de trabalho (onde estão os xlsx de teste).
engine_mod.BASE_DIR = WORK
cortex_mod.BASE_DIR = WORK

# Token: vem do ambiente (ou .env local). .env mantém DUMMY propositalmente.
TOKEN_TELEGRAM = cortex_mod._ler_env("TELEGRAM_TOKEN", WORK) or cortex_mod._ler_env("TELEGRAM_TOKEN", PARENT)
CHAT_ID_ALERTAS = cortex_mod._ler_env("TELEGRAM_CHAT_ID", WORK) or cortex_mod._ler_env("TELEGRAM_CHAT_ID", PARENT)

if not TOKEN_TELEGRAM:
    raise RuntimeError(
        "TELEGRAM_TOKEN não encontrado no ambiente nem no .env. "
        "Configure-o antes de iniciar o bot (nunca deixe o token hardcoded)."
    )

import telebot

bot = telebot.TeleBot(TOKEN_TELEGRAM)


# ----------------------------------------------------------------------
# Orquestração
# ----------------------------------------------------------------------
def _normalizar_aamm(texto_bruto: str) -> str:
    """
    Converte uma entrada de período para AAMM (%y%m).

    Aceita:
      - AAMM (ex: 2608)  -> direto
      - MMAA (ex: 0826)  -> convertido (mês=0826[:2], ano=0826[2:])
      - vazio/None       -> assume o período corrente (datetime.now)

    Retorna sempre no formato %y%m.
    """
    if not texto_bruto:
        return datetime.now().strftime("%y%m")
    t = texto_bruto.strip()
    # Se vier como MMAA (mês primeiro, ex: 0826), transpõe para AAMM.
    # Heurística: se os 2 primeiros dígitos forem > 12, é AAMM; senão MMAA.
    if re.fullmatch(r"\d{4}", t):
        mm, aa = t[:2], t[2:]
        if int(mm) > 12:
            return t  # já estava em AAMM
        return aa + mm  # vira AAMM
    return t


def _parse_aamm(texto: str):
    """
    Extrai o período do texto do comando (aceita AAMM ou MMAA).
    Se não houver número de 4 dígitos, assume o período corrente (hoje).
    """
    m = re.search(r"\b(\d{4})\b", texto or "")
    if m:
        return _normalizar_aamm(m.group(1))
    return _normalizar_aamm("")


async def _rodar(aamm: str = None, limite: int = None) -> dict:
    """Executa a reconcilação assíncrona (cortex -> engine -> balancete)."""
    engine = async_recon_mod.AsyncReconciliationEngine()
    resultado = await engine.run_reconciliation(aamm=aamm, limite=limite)
    return resultado


def _resumo(resultado: dict) -> str:
    """Monta a mensagem de resumo para o Telegram."""
    det = (resultado or {}).get("details") or {}
    cortex = det.get("cortex_padroeira", {})
    eng = det.get("engine_consolidacao", {})
    bal = det.get("motor_balancete", {})

    aamms_eng = (eng.get("aamms") or []) + (bal.get("aamms") or [])
    aamms_unicos = sorted(set(aamms_eng))

    linhas = []
    linhas.append("🤖 Reconciliação Padroeira — Resumo")
    if aamms_unicos:
        linhas.append(f"📅 Período(s): {', '.join(aamms_unicos)}")
    linhas.append(f"• Córtex (preflight): {cortex.get('status', '?')}"
                  + (f" — {cortex.get('error')}" if cortex.get("error") else ""))
    linhas.append(f"• Engine (Diário): {eng.get('status', '?')}")
    linhas.append(f"• Balancete (Pad): {bal.get('status', '?')}")

    extr = (cortex.get("details") or {}).get("extração_pendentes") or {}
    if extr:
        ok = sum(1 for v in extr.values() if v)
        linhas.append(f"• Fechamentos Saurus extraídos: {ok}/{len(extr)}")

    overall = (resultado or {}).get("status", "error")
    linhas.append(f"✅ Status geral: {overall}")
    return "\n".join(linhas)


def _fmt_brl(v: float) -> str:
    """Formata um valor float como R$ com separadores pt-BR (vírgula decimal)."""
    if v is None:
        return "—"
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _ler_faturamento_do_dia() -> dict:
    """
    Lê o FATURAMENTO DO DIA (linha 37 "Real" do Diário) da coluna correspondente
    à data de hoje em Movto_diario.AAMM.xlsx (AAMM corrente).

    Retorna um dict com a chave 'erro' (mensagem amigável) em caso de falha, ou:
        {
          'data': 'DD/MM/AAAA', 'aamm': 'AAMM',
          'real': float|None, 'sistema': float|None, 'sangria': float|None,
          'msg': str,            # mensagem pronta para o Telegram
        }
    O campo 'msg' já traz o texto formatado para a conferência de caixa.
    """
    hoje = datetime.now()
    aamm = hoje.strftime("%y%m")
    data_hoje = hoje.date()
    data_str = hoje.strftime("%d/%m/%Y")
    caminho = os.path.join(WORK, f"Movto_diario.{aamm}.xlsx")

    if not os.path.exists(caminho):
        msg = (f"📊 Faturamento do dia {data_str}\n"
               f"⚠️ Diário do período {aamm} não encontrado ({os.path.basename(caminho)}). "
               f"Rode /finalizar primeiro.")
        return {"erro": msg, "data": data_str, "aamm": aamm,
                "real": None, "sistema": None, "sangria": None, "msg": msg}

    try:
        wb = load_workbook(caminho, data_only=True)
        ws = wb.active
        col_dia = None
        for c in range(2, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if isinstance(v, datetime) and v.date() == data_hoje:
                col_dia = c
                break
        if not col_dia:
            wb.close()
            msg = (f"📊 Faturamento do dia {data_str}\n"
                   f"⚠️ Coluna do dia de hoje não encontrada no Diário {aamm}. "
                   f"O dia ainda não foi processado.")
            return {"erro": msg, "data": data_str, "aamm": aamm,
                    "real": None, "sistema": None, "sangria": None, "msg": msg}

        def _val(linha):
            v = ws.cell(row=linha, column=col_dia).value
            try:
                return float(v) if v not in (None, "") else None
            except (ValueError, TypeError):
                return None
        real = _val(37)
        sistema = _val(38)
        sangria = _val(42)
        wb.close()

        msg = (
            f"📊 Faturamento do dia {data_str}\n"
            f"💰 Real (Caixa): {_fmt_brl(real)}\n"
            f"🧮 Sistema: {_fmt_brl(sistema)}\n"
            f"🔻 Sangria: {_fmt_brl(sangria)}"
        )
        return {"erro": None, "data": data_str, "aamm": aamm,
                "real": real, "sistema": sistema, "sangria": sangria, "msg": msg}
    except Exception as e:
        msg = f"📊 Faturamento do dia {data_str}\n⚠️ Erro ao ler o Diário: {e}"
        return {"erro": msg, "data": data_str, "aamm": aamm,
                "real": None, "sistema": None, "sangria": None, "msg": msg}


def _salvar_historico_faturamento(registro: dict) -> bool:
    """
    Persiste o faturamento do dia no histórico (JSON) para uso futuro.

    O registro é indexado pela data (DD/MM/AAAA) e sobrescreve o do mesmo dia.
    Retorna True se salvou com sucesso.
    """
    if registro.get("erro"):
        return False
    try:
        historico = {}
        if os.path.exists(HIST_FILE):
            with open(HIST_FILE, "r", encoding="utf-8") as f:
                historico = json.load(f)
        historico[registro["data"]] = {
            "aamm": registro["aamm"],
            "real": registro["real"],
            "sistema": registro["sistema"],
            "sangria": registro["sangria"],
            "salvo_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        with open(HIST_FILE, "w", encoding="utf-8") as f:
            json.dump(historico, f, ensure_ascii=False, indent=2)
        logger.info(f"[HISTORICO] Faturamento de {registro['data']} salvo em {HIST_FILE}")
        return True
    except Exception:
        logger.exception("[HISTORICO] Falha ao salvar faturamento no histórico")
        return False


# ----------------------------------------------------------------------
# Handlers Telegram
# ----------------------------------------------------------------------
@bot.message_handler(commands=['finalizar', 'reconciliar', 'fechar'])
def cmd_finalizar(message):
    """
    Handler único dos comandos de operação.

    /fechar            -> PRIMEIRO comando do operador. Puxa o faturamento do dia,
                          lê para a CONFERÊNCIA DE CAIXA e SALVA NO HISTÓRICO.
                          Apenas leitura + gravação de histórico (não roda pipeline).
    /finalizar [MMAA]  -> CONCLUI o preenchimento e o transporte de dados
                          (Cortex -> Engine -> Balancete Pad). Sem MMAA, assume o
                          dia de hoje; com MMAA, varredura completa do período.
    /reconciliar [MMAA]-> alias de compatibilidade de /finalizar.
    """
    texto = message.text or ""
    chat_id = message.chat.id
    comando = (texto.split()[0].lstrip("/").lower() if texto.split() else "")

    logger.info("[TELEGRAM] Comando recebido do usuário.")

    # /fechar -> puxa faturamento, lê p/ conferência de caixa, salva no histórico.
    if comando == "fechar" and not re.search(r"\b\d{4}\b", texto):
        try:
            reg = _ler_faturamento_do_dia()
            bot.send_message(chat_id, reg["msg"])
            if reg.get("erro"):
                logger.info("[TELEGRAM] Faturamento do dia enviado ao usuário (com aviso).")
            else:
                salvou = _salvar_historico_faturamento(reg)
                logger.info(
                    f"[TELEGRAM] Faturamento do dia enviado e "
                    f"{'salvo no histórico' if salvou else 'NÃO salvo (erro de histórico)'}."
                )
        except Exception as e:
            logger.exception("[TELEGRAM] Erro ao enviar faturamento do dia")
            bot.send_message(chat_id, f"⚠️ Erro ao obter faturamento: {e}")
        return

    aamm = _parse_aamm(texto)
    escopo = "dia de hoje (AAMM corrente)" if not re.search(r"\b\d{4}\b", texto) else f"período {aamm} (varredura completa)"

    bot.reply_to(message, "🤖 Processando reconciliação Padroeira...\n"
                          f"Escopo: {escopo}")
    try:
        resultado = asyncio.run(_rodar(aamm=aamm))
        resumo = _resumo(resultado)
        bot.send_message(chat_id, resumo)
        logger.info("[TELEGRAM] Mensagem de resumo enviada ao usuário.")
    except Exception as e:
        logger.exception("[TELEGRAM] Erro ao executar reconciliação")
        try:
            bot.send_message(chat_id, f"⚠️ Erro na reconciliação: {e}")
        except Exception:
            pass


@bot.message_handler(commands=['amostra'])
def cmd_amostra(message):
    """/amostra [N] [AAMM|MMAA] — processa apenas N datas pendentes (teste e2e)."""
    texto = message.text or ""
    logger.info("[TELEGRAM] Comando recebido do usuário.")
    partes = texto.split()
    limite = 3
    aamm = None
    for p in partes[1:]:
        if p.isdigit() and int(p) < 100:
            limite = int(p)
        elif re.fullmatch(r"\d{4}", p):
            aamm = _normalizar_aamm(p)
    chat_id = message.chat.id

    bot.reply_to(message, f"🤖 Amostra de {limite} data(s) pendente(s) "
                          f"(período: {aamm or 'hoje/backlog'})...")
    try:
        resultado = asyncio.run(_rodar(aamm=aamm, limite=limite))
        resumo = _resumo(resultado)
        bot.send_message(chat_id, resumo)
        logger.info("[TELEGRAM] Mensagem de resumo enviada ao usuário.")
    except Exception as e:
        logger.exception("[TELEGRAM] Erro ao executar amostra")
        try:
            bot.send_message(chat_id, f"⚠️ Erro na amostra: {e}")
        except Exception:
            pass


@bot.message_handler(commands=['start', 'help'])
def cmd_help(message):
    bot.reply_to(
        message,
        "Comandos disponíveis:\n"
        "/fechar — puxa o FATURAMENTO DO DIA (linha 37 do Diário), lê para a "
        "conferência de caixa e SALVA NO HISTÓRICO. Use este PRIMEIRO.\n"
        "/finalizar [MMAA] — CONCLUI o preenchimento e o transporte de dados "
        "(Cortex -> Engine -> Balancete). Sem data, processa o DIA DE HOJE.\n"
        "/reconciliar [AAMM] — alias de /finalizar.\n"
        "/amostra [N] [MMAA] — roda N datas pendentes (teste e2e).\n"
    )


# ----------------------------------------------------------------------
# Loop principal — escuta continuamente.
# ----------------------------------------------------------------------
if __name__ == "__main__":
    logger.info("=" * 70)
    logger.info("Bot de Reconciliação Padroeira iniciado (modo escuta).")
    logger.info(f"Logs em tempo real: {LOGFILE}")
    logger.info("Comandos: /fechar (faturamento do dia -> historico)  |  /finalizar [MMAA]  |  /reconciliar [AAMM]  |  /amostra [N] [MMAA]")
    logger.info("=" * 70)
    bot.infinity_polling()
