#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Engine de Consolidação - FASE 2: Transposição Diário -> Balancete (Async Version)
Adaptado para Async Reconciliation Architecture V2
Correção 3 v2 (SANDBOX): limpeza dinâmica da zona de dias + detecção da âncora
na coluna D + realinhamento de Particip./Projeção após insert_rows.
"""

import os
import re
import shutil
import glob
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import logging
from typing import Dict, Any, Optional

from calendario_padroeira import (
    dia_eh_fechado,
    dias_do_mes,
    sinal_auxiliar,
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger("MotorBalanceteAsync")

# Todos os caminhos são resolvidos a partir do diretório deste script.
# Elimina referências hardcoded à nuvem (Box).
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def _encontrar_arquivo_base(prefixo: str) -> Optional[str]:
    """Localiza o arquivo base mais recente de um prefixo no diretório local."""
    candidatos = glob.glob(os.path.join(BASE_DIR, f"{prefixo}*.xlsx"))
    candidatos = [c for c in candidatos if "template_" not in os.path.basename(c)]
    if not candidatos:
        return None
    candidatos.sort(key=os.path.getmtime, reverse=True)
    return candidatos[0]

# ==============================================================================
# DICIONÁRIO DE TRANSPOSIÇÃO (Matriz Cruzada)
# Chave: Coluna Destino no Balancete | Valor: Linha Origem no Movto Diário
# ==============================================================================
MAPA_DIARIO_PAD = {
    'B': 3,   # Peso Buf.
    'C': 4,   # Peso Sob.
    'E': 6,   # Dinheiro
    'F': 10,  # MasterCard (Redecard crédito)
    'G': 11,  # RedeShop (Redecard debito)
    'H': 12,  # Visa Cred (Cielo Visa Credito)
    'I': 13,  # Visa Deb (Cielo Visa Débito)
    'J': 7,   # Cheques
    'K': 24,  # Tickets (Linha de Total dos tickets)
    'L': 27,  # Cvs/E (Pix / conta corrente)
    'N': 32,  # Pend.Dia
    'O': 33,  # Pend.Pg.
    'P': 36,  # Serviço
    'Q': 37,  # Mov/Dia (Real)
    'R': 42   # Sangria (Linha 42 do Diário; regra do CLAUDE.md: espelhada no Balancete)
}

# Rótulos estruturais que delimitam o fim da zona de dias no balancete.
ROTULOS_ESTRUTURAIS = ("Particip.", "Projeção", "Encargos")

class MotorBalanceteAsync:
    """
    Async version of Motor Balancete for integration with Async Reconciliation V2
    """

    def __init__(self, aamm: str = None):
        self.aamm = aamm or datetime.now().strftime("%y%m")
        self.status = {
            "file_check": None,
            "data_extraction": None,
            "transposition": None,
            "save": None
        }
        self.stats = {
            "days_processed": 0,
            "lines_modified": 0,
            "new_days_added": 0
        }

    def _limpar_dados_pad(self, caminho: str) -> None:
        """
        Limpa os dados das linhas de dias de um Balancete copiado,
        incluindo os números dos dias (coluna A). Preserva cabeçalhos e
        fórmulas estruturais (ex.: coluna Q = SUM(B:P), linha de totais).
        """
        wb = load_workbook(caminho, data_only=False)
        nome_aba = "Movimento Diario"
        ws = wb[nome_aba] if nome_aba in wb.sheetnames else wb.active

        # Zona de dias detectada dinamicamente: linhas 2..(primeiro rótulo estrutural - 1).
        # Rótulos ('Particip.', 'Projeção', 'Encargos') delimitam o fim da zona.
        fim_zona = 30
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and v.strip() in ROTULOS_ESTRUTURAIS:
                fim_zona = r - 1
                break
        # Linha de totais (âncora): 1ª linha com total estrutural =SUM(D2:D...) na coluna D.
        linha_totais = 29
        for r in range(2, fim_zona + 1):
            v = ws.cell(row=r, column=4).value
            if isinstance(v, str) and v.upper().startswith("=SUM(D"):
                linha_totais = r
                break
        for r in range(2, fim_zona + 1):
            for c in range(1, 18):  # 1..17 (A..Q); coluna R (Sangria) é tratada abaixo
                if r == linha_totais and c == 4:  # preserva total estrutural da coluna D
                    continue
                ws.cell(row=r, column=c).value = None
            # Limpa Q de slot de dia (=SUM(Br:Pr)); Q da âncora é reconstruído no injetor.
            q = ws.cell(row=r, column=17).value
            if isinstance(q, str) and q.upper().startswith("=SUM(B"):
                ws.cell(row=r, column=17).value = None
            # Limpa a coluna R (Sangria) de dias anteriores para garantir idempotência
            # (a Sangria é valor fixo transportado do Diário; não é fórmula de subtotal).
            ws.cell(row=r, column=18).value = None
        wb.save(caminho)

    def _garantir_planilha_pad(self) -> str:
        """
        Retorna o caminho de Pad{AAMM}.xlsx, criando-o se necessário.
        Estratégia: template_Pad.xlsx se existir; senão o Pad.*.xlsx mais recente
        (com limpeza dos dados dos dias, mantendo estrutura e fórmulas).
        """
        pad_path = os.path.join(BASE_DIR, f"Pad{self.aamm}.xlsx")
        if os.path.exists(pad_path):
            return pad_path

        template = os.path.join(BASE_DIR, "template_Pad.xlsx")
        base = None
        usar_limpeza = False
        if os.path.exists(template):
            base = template
        else:
            base = _encontrar_arquivo_base("Pad")
            usar_limpeza = True

        if not base:
            raise FileNotFoundError(
                f"Nenhum modelo (template_Pad.xlsx) nem base (Pad*.xlsx) "
                f"disponível para criar {os.path.basename(pad_path)}"
            )

        shutil.copy2(base, pad_path)
        if usar_limpeza:
            self._limpar_dados_pad(pad_path)
        logger.info(f"[+] Planilha {os.path.basename(pad_path)} criada a partir de {os.path.basename(base)}")
        return pad_path

    def normalizar_dia(self, valor_data) -> Optional[int]:
        """Extrai apenas o número do dia (inteiro) da data do Diário"""
        if isinstance(valor_data, datetime):
            return int(valor_data.day)
        if isinstance(valor_data, str):
            try:
                return int(datetime.strptime(valor_data.strip().split()[0], "%d/%m/%Y").day)
            except:
                try:
                    return int(datetime.strptime(valor_data.strip().split()[0], "%Y-%m-%d").day)
                except:
                    pass
        return None

    def _encontrar_linha_ancora(self, ws) -> int:
        """
        Localiza a linha de totais do balancete.
        (a) 1ª linha com total estrutural =SUM(D2:D...) na coluna D (referenciado
            por Particip./Projeção). Fallback (b): linha de 'Particip.' - 2;
            fallback final 29.
        """
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=4).value
            if isinstance(v, str) and v.upper().startswith("=SUM(D"):
                return r
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and str(v).strip().startswith("Particip"):
                return r - 2
        logger.warning("Detecção de linha de totais caiu no fallback 29 — verificar estrutura do Pad.")
        return 29

    def _encontrar_linhas_estruturais(self, ws, base: int) -> Dict[str, int]:
        """
        Localiza as linhas de Particip./Projeção/Encargos a partir da âncora.
        Usa o rótulo da coluna A quando disponível; senão offsets relativos à âncora.
        """
        linhas = {}
        for r in range(base + 1, min(base + 6, ws.max_row + 1)):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str):
                t = v.strip()
                for rotulo in ROTULOS_ESTRUTURAIS:
                    if t.startswith(rotulo[:5]):
                        linhas[rotulo] = r
        # Fallbacks relativos à âncora (padrão do template: +2 Particip., +3 Projeção, +4 Encargos)
        if "Particip." not in linhas:
            linhas["Particip."] = base + 2
        if "Projeção" not in linhas:
            linhas["Projeção"] = base + 3
        if "Encargos" not in linhas:
            linhas["Encargos"] = base + 4
        return linhas

    def _realinhar_fórmulas_estruturais(self, ws, linha_totais: int, linha_totais_antiga: Optional[int]) -> None:
        """
        Após insert_rows, openpyxl NÃO ajusta referências de célula em fórmulas.
        Reaponta Particip./Projeção/Encargos para a nova linha de totais.
        """
        linhas = self._encontrar_linhas_estruturais(ws, linha_totais)
        particip, projecao = linhas["Particip."], linhas["Projeção"]

        # Particip.: D{particip} = =D{totais}/Q{totais}
        ws.cell(row=particip, column=4).value = f"=D{linha_totais}/Q{linha_totais}"

        # Projeção: Q{proj} referenciando Q{totais}; D{proj} = =Q{proj}*D{particip}
        q_formula = ws.cell(row=projecao, column=17).value
        if isinstance(q_formula, str) and "Q" in q_formula:
            if linha_totais_antiga:
                q_formula = re.sub(rf"\bQ{linha_totais_antiga}\b", f"Q{linha_totais}", q_formula)
            ws.cell(row=projecao, column=17).value = q_formula
        ws.cell(row=projecao, column=4).value = f"=Q{projecao}*D{particip}"

        logger.info(
            f"    [*] Fórmulas estruturais reapontadas: "
            f"Particip. D{particip}=D{linha_totais}/Q{linha_totais}, "
            f"Projeção D{projecao}=Q{projecao}*D{particip}, Q{projecao}=Q{linha_totais}/..."
        )

    def _validar_dias_ausentes(self, dias_capturados: set) -> Dict[str, list]:
        """
        Regra de negócio (fcd26ad3 simplificada): "dia fechado" = AUSÊNCIA de
        registro no caixa (sem fechamento_caixa_{data}.txt E sem entrada no
        Movto_cx2). Calendário/feriado não decide; é só sinal auxiliar de log.
        Classifica os dias do mês que NÃO têm movimento no Diário em:
          - 'fechados': sem registro no caixa — ausência ESPERADA, não bug.
          - 'em_aberto': com registro no caixa mas sem Diário — possível bug real.
        Retorna {'fechados': [...], 'em_aberto': [...]}.
        """
        aamm = self.aamm
        todos = set(range(1, dias_do_mes(aamm) + 1))
        ausentes = sorted(todos - dias_capturados)
        fechados = []
        em_aberto = []
        for dia in ausentes:
            if dia_eh_fechado(BASE_DIR, aamm, dia):
                fechados.append(dia)
            else:
                em_aberto.append(dia)
        return {"fechados": fechados, "em_aberto": em_aberto}

    def injetar_balancete(self) -> Dict[str, Any]:
        """
        Execute the balance sheet injection process
        """
        try:
            diario_path = os.path.join(BASE_DIR, f"Movto_diario.{self.aamm}.xlsx")

            # O Diário é fonte obrigatória (gerado pelo Engine). O Pad é auto-criado.
            if not os.path.exists(diario_path):
                error_msg = f"Movto_diario.{self.aamm}.xlsx não encontrado. Execute o Engine de Consolidação antes."
                logger.error(error_msg)
                self.status["file_check"] = f"error: {error_msg}"
                return {"status": "error", "error": error_msg}

            pad_path = self._garantir_planilha_pad()

            self.status["file_check"] = "success"
            logger.info("[*] Iniciando Fase 2: Transposição de Matriz (Diário -> Balancete Pad)...")

            # 1. Carrega o Diário para LEITURA dos valores finais calculados
            wb_diario = load_workbook(diario_path, data_only=True)
            ws_diario = wb_diario.active

            # Extrai a carga útil de todos os dias disponíveis no Diário
            carga_por_dia = {}
            for col in range(2, ws_diario.max_column + 1):
                v_data = ws_diario.cell(row=1, column=col).value
                if v_data:
                    dia_int = self.normalizar_dia(v_data)
                    if dia_int:
                        carga_por_dia[dia_int] = {}
                        # Coleta os valores conforme o mapa
                        for col_pad, linha_diario in MAPA_DIARIO_PAD.items():
                            valor = ws_diario.cell(row=linha_diario, column=col).value
                            # Substitui None por 0 para não quebrar fórmulas matemáticas no Balancete
                            carga_por_dia[dia_int][col_pad] = valor if valor is not None else 0.0

            if not carga_por_dia:
                error_msg = "Nenhum dia válido encontrado no Diário Mensal"
                logger.error(error_msg)
                self.status["data_extraction"] = f"error: {error_msg}"
                return {"status": "error", "error": error_msg}

            logger.info(f"[*] Carga extraída do Diário! Dias capturados: {sorted(list(carga_por_dia.keys()))}")
            self.status["data_extraction"] = "success"
            self.stats["days_processed"] = len(carga_por_dia)

            # Regra de negócio (fcd26ad3 simplificada): dia fechado = AUSÊNCIA de
            # registro no caixa. Dias sem registro são ESPERADOS, não bug. Dias com
            # registro no caixa mas sem Diário são candidatos a dado faltante real
            # — sinalizar sem abortar.
            classificacao = self._validar_dias_ausentes(set(carga_por_dia.keys()))
            if classificacao["fechados"]:
                anotacao = ", ".join(
                    f"{d}({sinal_auxiliar(self.aamm, d)})"
                    for d in classificacao["fechados"]
                )
                logger.info(
                    f"[calendário] Dias SEM registro no caixa (fechado esperado, "
                    f"não bug): {anotacao}"
                )
            if classificacao["em_aberto"]:
                logger.warning(
                    f"[calendário] Dias com registro no caixa mas SEM movimento no "
                    f"Diário — verificar dado faltante real: "
                    f"{classificacao['em_aberto']}"
                )
            self.stats["dias_fechados"] = len(classificacao["fechados"])
            self.stats["dias_uteis_ausentes"] = len(classificacao["em_aberto"])

            # 2. Carrega o Balancete (Pad) em modo ESCRITA na aba "Movimento Diario"
            wb_pad = load_workbook(pad_path, data_only=False)

            # Garante que estamos escrevendo na aba certa, independente de ser a ativa
            nome_aba = "Movimento Diario"
            if nome_aba in wb_pad.sheetnames:
                ws_pad = wb_pad[nome_aba]
            else:
                ws_pad = wb_pad.active
                logger.warning(f"Aba '{nome_aba}' não encontrada. Usando a aba ativa: {ws_pad.title}")

            # 3. Varredura e Sobreposição: dias em ordem crescente a partir da linha 2,
            #    garantindo espaço até o dia 31 (insere linhas antes da linha de totais se necessário).
            linhas_modificadas = 0
            novos_dias = 0
            dias_ordenados = sorted(carga_por_dia.keys())

            # (a) Linha de totais (âncora): 1ª linha com total estrutural =SUM(D2:D...) na
            #     coluna D (referenciado por Particip./Projeção).
            linha_totais = self._encontrar_linha_ancora(ws_pad)
            linha_totais_antiga = linha_totais

            # (b) Garante espaço: cada dia ocupa uma linha a partir da linha 2.
            linha_ultimo_dia = 1 + len(dias_ordenados)
            if linha_ultimo_dia >= linha_totais:
                n_inserir = linha_ultimo_dia - linha_totais + 1
                ws_pad.insert_rows(linha_totais, n_inserir)
                logger.info(
                    f"    [+] Inseridas {n_inserir} linha(s) antes da linha {linha_totais} "
                    f"para garantir espaço até o dia {len(dias_ordenados)}."
                )
                nova_linha_totais = linha_totais + n_inserir
                for c in range(2, 18):
                    letra = get_column_letter(c)
                    ws_pad[f"{letra}{nova_linha_totais}"] = f"=SUM({letra}2:{letra}{linha_ultimo_dia})"
                linha_totais = nova_linha_totais  # atualiza a âncora após deslocamento
                # (b1) Reaponta Particip./Projeção/Encargos para a nova âncora.
                self._realinhar_fórmulas_estruturais(ws_pad, linha_totais, linha_totais_antiga)

            # (c) Zona de dias: linhas 2..(primeiro rótulo estrutural - 1). Detectada APÓS
            #     qualquer inserção, cobrindo também resíduos abaixo da âncora (ex.: linhas
            #     29-30 com dias 7/21 herdados da base Pad2608).
            fim_zona_dias = linha_totais
            for r in range(2, ws_pad.max_row + 1):
                v = ws_pad.cell(row=r, column=1).value
                if isinstance(v, str) and str(v).strip() in ROTULOS_ESTRUTURAIS:
                    fim_zona_dias = r - 1
                    break
            if fim_zona_dias < linha_totais:
                fim_zona_dias = linha_totais

            # (d) Limpa TODA a zona de dias (inclui fantasmas 29/30); preserva na âncora
            #     apenas o total estrutural da coluna D. Q é limpo e reconstruído abaixo.
            for r in range(2, fim_zona_dias + 1):
                for c in range(1, 17):
                    if r == linha_totais and c == 4:
                        continue
                    ws_pad.cell(row=r, column=c).value = None
                ws_pad.cell(row=r, column=17).value = None

            # (e) Grava os dias em ordem crescente, um por linha a partir da linha 2.
            # Garante o cabeçalho da coluna R (Sangria) — regra do CLAUDE.md: sangria
            # na Linha 42 do Diário deve ser espelhada no Balancete (coluna R).
            ws_pad[f"R1"] = "Sangria"
            for i, dia in enumerate(dias_ordenados):
                linha_destino = 2 + i
                if ws_pad.cell(row=linha_destino, column=1).value is None:
                    novos_dias += 1
                ws_pad.cell(row=linha_destino, column=1, value=dia)
                valores = carga_por_dia[dia]
                for letra_coluna, valor in valores.items():
                    if letra_coluna == 'Q':
                        continue
                    ws_pad[f"{letra_coluna}{linha_destino}"].value = valor
                ws_pad[f'Q{linha_destino}'].value = f'=SUM(B{linha_destino}:P{linha_destino})'
                linhas_modificadas += 1

            # (f) Reconstrói as fórmulas de agregação da linha de totais (D/Q e demais
            #     colunas B..P) sobre os dias efetivamente gravados, para que
            #     Particip. (D31=D29/Q29) e Projeção (Q32=Q29/...) apontem para totais
            #     corretos. Remove a fórmula de subtotal espúria do slot ocioso
            #     (linha_totais-1) quando não usado.
            ws_pad[f"D{linha_totais}"] = f"=SUM(D2:D{linha_ultimo_dia})"
            ws_pad[f"Q{linha_totais}"] = f"=SUM(Q2:Q{linha_ultimo_dia})"
            ws_pad[f"R{linha_totais}"] = f"=SUM(R2:R{linha_ultimo_dia})"  # Sangria (espelhada do Diário)
            for c in range(2, 19):  # 2..18 (B..R) — inclui a coluna R (Sangria)
                letra = get_column_letter(c)
                ws_pad[f"{letra}{linha_totais}"] = f"=SUM({letra}2:{letra}{linha_ultimo_dia})"
            linha_slot_final = linha_totais - 1
            if linha_ultimo_dia < linha_slot_final:
                ws_pad[f"Q{linha_slot_final}"] = (
                    f"=SUM(D{linha_slot_final}:N{linha_slot_final})"
                    f"-P{linha_slot_final}-O{linha_slot_final}"
                )

            self.stats["lines_modified"] = linhas_modificadas
            self.stats["new_days_added"] = novos_dias
            self.status["transposition"] = "success"

            # 4. Salva o Balancete
            if linhas_modificadas > 0:
                try:
                    wb_pad.save(pad_path)
                    logger.info(f"\n[+] Fase 2 Concluída! {linhas_modificadas} dias sobrepostos com sucesso no Balancete (Pad{self.aamm}.xlsx).")
                    self.status["save"] = "success"
                    return {
                        "status": "success",
                        "stats": self.stats,
                        "details": self.status
                    }
                except Exception as e:
                    error_msg = f"Erro ao salvar o Balancete: {str(e)}"
                    logger.error(error_msg)
                    self.status["save"] = f"error: {str(e)}"
                    return {"status": "error", "error": error_msg}
            else:
                logger.info("Nenhuma modificação necessária no Balancete.")
                self.status["save"] = "skipped"
                return {
                    "status": "success",
                    "stats": self.stats,
                    "details": self.status
                }

        except Exception as e:
            error_msg = f"Erro inesperado no motor de balancete: {str(e)}"
            logger.error(error_msg)
            return {"status": "error", "error": error_msg}

    def get_status(self) -> Dict[str, Any]:
        """Return the current status of the motor"""
        return {
            "status": self.status,
            "stats": self.stats,
            "aamm": self.aamm
        }

if __name__ == "__main__":
    logger.info("[*] =================================================================")
    logger.info("[*] PHASE 2: MOTOR DE TRANSPOSIÇÃO E SOBREPOSIÇÃO AGRESSIVA - Async Version")
    logger.info("[*] =================================================================")

    motor = MotorBalanceteAsync()
    result = motor.injetar_balancete()

    if result["status"] == "success":
        logger.info("Motor de balancete executado com sucesso!")
        logger.info(f"Estatísticas: {motor.get_status()['stats']}")
    else:
        logger.error(f"Erro no motor de balancete: {result['error']}")
