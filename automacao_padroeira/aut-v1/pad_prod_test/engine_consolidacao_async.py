#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Engine de Consolidação - Mestre Padroeira (Async Version)
Motor Unificado adaptado para Async Reconciliation Architecture V2
"""

import os
import shutil
import glob
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import logging
from typing import Dict, Any, List, Optional

from backup_padroeira import snapshot_arquivo, registrar_divergencia, registrar_checkup_dia

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger("EngineConsolidacaoAsync")

# Todos os caminhos são resolvidos a partir do diretório deste script.
# Isso garante portabilidade e elimina referências hardcoded à nuvem (Box).
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Nada anterior a isso é tocado, mesmo numa reexecução "do zero" —
# meses anteriores já estão fechados/conferidos manualmente.
DATA_MINIMA_PROCESSAMENTO = datetime(2026, 6, 1).date()

# ATENÇÃO: a linha 37 representa o total do caixa do dia (slips + dinheiro)
# no Movto_diario.xlsx. Atualizado conforme confirmação do usuário em ago/2026.
LINHA_TOTAL_CAIXA = 37

def _encontrar_arquivo_base(prefixo: str, ignorar_templates: bool = True) -> Optional[str]:
    """
    Localiza o arquivo base mais recente de um prefixo no diretório local.
    Ex: prefixo 'Movto_diario' -> Movto_diario.*.xlsx (mais recente por mtime).
    """
    padroes = [f"{prefixo}*.xlsx"]
    if ignorar_templates:
        padroes.append(f"template_{prefixo}*.xlsx")
    candidatos = []
    for p in padroes:
        candidatos.extend(glob.glob(os.path.join(BASE_DIR, p)))
    # Exclui arquivos de template da busca de base (template é prioridade separada)
    candidatos = [c for c in candidatos if "template_" not in os.path.basename(c)]
    if not candidatos:
        return None
    candidatos.sort(key=os.path.getmtime, reverse=True)
    return candidatos[0]

class EngineConsolidacaoAsync:
    """
    Async version of Engine de Consolidação for integration with Async Reconciliation V2
    """

    def __init__(self, aamm: str = None):
        self.aamm = aamm or datetime.now().strftime("%y%m")
        self.status = {
            "file_check": None,
            "expansion": None,
            "injection": None,
            "save": None
        }
        self.stats = {
            "new_columns": 0,
            "columns_updated": 0,
            "cells_modified": 0
        }

    def _limpar_colunas_diario(self, caminho: str) -> None:
        """
        Limpa os dados das colunas de dias de uma planilha Diário copiada,
        incluindo o cabeçalho (linha 1) das colunas B em diante — o engine
        regenera o calendário do mês-alvo na expansão. Preserva os rótulos
        da coluna A.
        As linhas 14/24/41 (controle e fórmulas por coluna) são limpas para que
        o motor as regenere corretamente para o novo período.
        """
        wb = load_workbook(caminho, data_only=False)
        ws = wb.active
        for col in range(2, ws.max_column + 1):
            if ws.cell(row=1, column=col).value is not None:
                for row in range(1, ws.max_row + 1):
                    ws.cell(row=row, column=col).value = None
        wb.save(caminho)

    def _garantir_planilha_diaria(self) -> str:
        """
        Retorna o caminho de Movto_diario.{AAMM}.xlsx, criando-o se necessário.
        Estratégia de criação (em ordem de prioridade):
          1. template_Movto_diario.xlsx (modelo limpo) se existir -> cópia direta.
          2. Arquivo Movto_diario.*.xlsx mais recente -> cópia + limpeza dos dados
             das colunas de dias (mantém cabeçalhos/rótulos).
        """
        caminho = os.path.join(BASE_DIR, f"Movto_diario.{self.aamm}.xlsx")
        if os.path.exists(caminho):
            return caminho

        template = os.path.join(BASE_DIR, "template_Movto_diario.xlsx")
        base = None
        usar_limpeza = False
        if os.path.exists(template):
            base = template
        else:
            base = _encontrar_arquivo_base("Movto_diario")
            usar_limpeza = True

        if not base:
            raise FileNotFoundError(
                f"Nenhum modelo (template_Movto_diario.xlsx) nem base "
                f"(Movto_diario.*.xlsx) disponível para criar {os.path.basename(caminho)}"
            )

        shutil.copy2(base, caminho)
        if usar_limpeza:
            self._limpar_colunas_diario(caminho)
        logger.info(f"[+] Planilha {os.path.basename(caminho)} criada a partir de {os.path.basename(base)}")
        return caminho

    def normalizar_data(self, valor) -> Optional[datetime.date]:
        """Retorna um objeto date para casamento seguro de colunas"""
        if isinstance(valor, datetime):
            return valor.date()
        if isinstance(valor, str):
            try:
                return datetime.strptime(valor.strip().split()[0], "%d/%m/%Y").date()
            except:
                try:
                    return datetime.strptime(valor.strip().split()[0], "%Y-%m-%d").date()
                except:
                    return None
        return None

    def executar_motor_unificado(
        self,
        dados_cortex: Optional[Dict[str, Dict[str, str]]] = None,
        bot=None,
        chat_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Execute the unified engine for data consolidation.
        Optionally receives `dados_cortex` (a dict keyed by ISO date strings such as
        "2026-06-01" with per-day Saurus metrics) to inject peso e cliente metrics
        into the upper rows of the Movto_diario spreadsheet.

        `bot`/`chat_id`: instância do telebot e o chat de destino, usados para
        alertar divergências de caixa acima de LIMITE_DIVERGENCIA (ver
        backup_padroeira.registrar_divergencia). Se omitidos, a divergência
        ainda é calculada e gravada no SQLite, só não dispara Telegram.
        """
        try:
            caminho_caixa2 = os.path.join(BASE_DIR, "Movto_cx2.xlsx")

            # O Caixa 2 é fonte obrigatória; a planilha do mês é auto-criada se faltar.
            if not os.path.exists(caminho_caixa2):
                error_msg = "Arquivo base Movto_cx2.xlsx não encontrado"
                logger.error(error_msg)
                self.status["file_check"] = f"error: {error_msg}"
                return {"status": "error", "error": error_msg}

            caminho_mensal = self._garantir_planilha_diaria()

            # Backup ANTES de qualquer escrita: snapshot do estado atual do
            # arquivo mensal (se já existia) e do Caixa 2 (fonte), guardados
            # no SQLite local em vez de mais uma planilha.
            snapshot_arquivo(caminho_mensal, aamm=self.aamm)
            snapshot_arquivo(caminho_caixa2, aamm=self.aamm)

            self.status["file_check"] = "success"
            logger.info("[*] Carregando matrizes na memória...")

            # Caixa 2 (Leitura dos dados da Sandra)
            wb_cx = load_workbook(caminho_caixa2, data_only=True)
            ws_cx = wb_cx.active

            # Diário Mensal (Escrita - preserva fórmulas)
            wb_me = load_workbook(caminho_mensal, data_only=False)
            ws_me = wb_me.active

            # --- Corrige rótulos estruturais (Sangria permanece na Linha 42) ---
            # A fórmula de Total de Caixa (=col37-col38) fica na Linha 40 (vide copy loop).
            # O rótulo "Sangria" fica estritamente na Linha 42, espelhado do Cx2.
            if str(ws_me.cell(row=40, column=1).value or "").strip().lower() == "sangria":
                ws_me.cell(row=40, column=1).value = None
            if str(ws_me.cell(row=41, column=1).value or "").strip().lower() == "total caixa":
                ws_me.cell(row=41, column=1).value = None
            if str(ws_me.cell(row=42, column=1).value or "").strip().lower() != "sangria":
                ws_me.cell(row=42, column=1, value="Sangria")

            # Diário Mensal (Leitura - checa resultados das fórmulas existentes)
            wb_me_lei = load_workbook(caminho_mensal, data_only=True)
            ws_me_lei = wb_me_lei.active

            # ==============================================================================
            # ETAPA 1: MAPEAMENTO E EXPANSÃO HORIZONTAL
            # ==============================================================================
            datas_no_mensal = {}
            colunas_pendentes_de_carga = []
            proxima_coluna_livre = 2

            # 1.1 Varre o Diário atual para entender o que já existe
            for col in range(2, ws_me_lei.max_column + 1):
                v_data = ws_me_lei.cell(row=1, column=col).value
                v_total = ws_me_lei.cell(row=24, column=col).value  # Linha de controle

                if v_data:
                    dt_norm = self.normalizar_data(v_data)
                    if dt_norm:
                        datas_no_mensal[dt_norm] = col
                        # Se o dia existe, mas a linha 24 está vazia/zerada, agenda para receber carga
                        if v_total is None or float(v_total) == 0:
                            colunas_pendentes_de_carga.append(col)

                # Identifica a borda da matriz para criar novas datas
                if ws_me_lei.cell(row=1, column=col).value is None and proxima_coluna_livre == 2:
                    proxima_coluna_livre = col

            if proxima_coluna_livre == 2:
                proxima_coluna_livre = ws_me_lei.max_column + 1

            # 1.2 Lê o Caixa 2 e separa as datas do mês ativo
            mes_alvo = int(self.aamm[2:])
            ano_alvo = 2000 + int(self.aamm[:2])
            mapa_cx_datas = {}
            datas_faltantes = []

            for col in range(2, ws_cx.max_column + 1):
                v_data_cx = ws_cx.cell(row=1, column=col).value
                if v_data_cx:
                    dt_cx_norm = self.normalizar_data(v_data_cx)
                    if dt_cx_norm and dt_cx_norm < DATA_MINIMA_PROCESSAMENTO:
                        continue  # fora de escopo: antes de jun/2026, não toca
                    if dt_cx_norm and dt_cx_norm.month == mes_alvo and dt_cx_norm.year == ano_alvo:
                        mapa_cx_datas[dt_cx_norm] = col
                        if dt_cx_norm not in datas_no_mensal and dt_cx_norm not in datas_faltantes:
                            datas_faltantes.append(dt_cx_norm)

            datas_faltantes.sort()

            # 1.3 Injeta as colunas novas na Linha 1 do Diário (se houver)
            if datas_faltantes:
                logger.info(f"[*] Expandindo o calendário: {len(datas_faltantes)} novos dias detectados.")
                col_atual = proxima_coluna_livre
                for nova_data in datas_faltantes:
                    dt_objeto = datetime(nova_data.year, nova_data.month, nova_data.day)
                    ws_me.cell(row=1, column=col_atual, value=dt_objeto)
                    ws_me.cell(row=1, column=col_atual).number_format = 'd-mmm-yy'

                    # Registra no mapa e já agenda a nova coluna para receber carga
                    datas_no_mensal[nova_data] = col_atual
                    colunas_pendentes_de_carga.append(col_atual)
                    col_atual += 1

                self.stats["new_columns"] = len(datas_faltantes)
            else:
                logger.info("[+] A matriz de calendário já está sincronizada.")

            self.status["expansion"] = "success"

            # Alvos da verificação pós-escrita (post-flight). Coletados durante a
            # injeção: guarda a coluna mensal, a data ISO e os dados do Cortex para
            # revalidar as células 3/4/5 após a gravação.
            alvos_verificacao: List[Dict[str, Any]] = []

            # ==============================================================================
            # ETAPA 2: ESPELHAMENTO VERTICAL (INJEÇÃO DE DADOS E FÓRMULAS)
            # ==============================================================================
            if not colunas_pendentes_de_carga:
                logger.info("[+] Paridade de faturamento total! Nenhuma coluna precisa de dados.")
                self.status["injection"] = "success"
            else:
                logger.info(f"\n[*] Iniciando injeção em {len(colunas_pendentes_de_carga)} colunas pendentes...")
                mudancas = 0

                # Filtra o mapa para garantir que temos o dado no Caixa 2
                dias_para_processar = []
                for dt, col_me in datas_no_mensal.items():
                    if col_me in colunas_pendentes_de_carga and dt in mapa_cx_datas:
                        dias_para_processar.append({
                            "data": dt,
                            "col_caixa": mapa_cx_datas[dt],
                            "col_mensal": col_me
                        })

                # Executa o paralelismo vertical
                for dia in dias_para_processar:
                    letra_col_me = get_column_letter(dia['col_mensal'])
                    logger.debug(f"    -> Transportando dados para {dia['data'].strftime('%d/%m/%Y')} (Coluna {letra_col_me})")

                    # ---- Injeção das métricas do Saurus (peso & clientes) ----
                    # Recupera os dados do Cortex para a data do dia processado.
                    data_iso = dia['data'].strftime("%Y-%m-%d")
                    dados_dia = dados_cortex.get(data_iso) if dados_cortex else None
                    if dados_dia:
                        self.stats["cells_modified"] = self.stats.get("cells_modified", 0)
                        try:
                            # Injeta Refeição Kg Equivalente na Linha 3
                            if "kg_eq_ref" in dados_dia and dados_dia['kg_eq_ref'] not in (None, ""):
                                ws_me.cell(row=3, column=dia['col_mensal'], value=float(dados_dia['kg_eq_ref']))
                                self.stats["cells_modified"] += 1

                            # Injeta Sobremesa/Doces Kg Equivalente na Linha 4
                            if "kg_eq_sob" in dados_dia and dados_dia['kg_eq_sob'] not in (None, ""):
                                ws_me.cell(row=4, column=dia['col_mensal'], value=float(dados_dia['kg_eq_sob']))
                                self.stats["cells_modified"] += 1

                            # Injeta Quantidade de Clientes na Linha 5
                            if "clientes" in dados_dia and dados_dia['clientes'] not in (None, ""):
                                ws_me.cell(row=5, column=dia['col_mensal'], value=int(dados_dia['clientes']))
                                self.stats["cells_modified"] += 1
                        except (ValueError, TypeError) as e:
                            logger.warning(f"[AVISO] Falha ao injetar métrica Saurus p/ {data_iso}: {e}")

                        # Registra o alvo para revalidar células 3/4/5 após a gravação.
                        alvos_verificacao.append({
                            "col_mensal": dia['col_mensal'],
                            "data_iso": data_iso,
                            "dados": dados_dia,
                        })

                    for linha in range(6, ws_cx.max_row + 1):
                        # Preserva fórmulas vitais
                        if linha == 14:
                            ws_me.cell(row=linha, column=dia['col_mensal'], value=f"=SUM({letra_col_me}10:{letra_col_me}13)")
                            mudancas += 1
                            continue
                        if linha == 40:
                            ws_me.cell(row=linha, column=dia['col_mensal'], value=f"={letra_col_me}37-{letra_col_me}38")
                            mudancas += 1
                            continue

                        # Injeta valor bruto
                        valor_real_cx = ws_cx.cell(row=linha, column=dia['col_caixa']).value
                        ws_me.cell(row=linha, column=dia['col_mensal'], value=valor_real_cx)
                        mudancas += 1

                self.stats["cells_modified"] = mudancas
                self.stats["columns_updated"] = len(colunas_pendentes_de_carga)
                logger.info(f"[+] Espelhamento concluído! {mudancas} células tratadas.")
                self.status["injection"] = "success"

            # ==============================================================================
            # ETAPA 2.5: INJEÇÃO IDEMPOTENTE DAS MÉTRICAS SAURUS (LINHAS 3/4/5)
            # ==============================================================================
            # Não depende das colunas "pendentes": percorre TODAS as colunas do mês-alvo
            # com data em dados_cortex e garante que Peso buffet (3) / Peso sobremesa (4) /
            # Clientes (5) estejam preenchidos. Idempotente — pode rodar quantas vezes for,
            # sempre reconciliando com a fonte (cache de fechamentos).
            metricas_saurus = 0
            if dados_cortex:
                for dt, col in datas_no_mensal.items():
                    if dt.month != mes_alvo or dt.year != ano_alvo:
                        continue
                    data_iso = dt.strftime("%Y-%m-%d")
                    dados_dia = dados_cortex.get(data_iso)
                    if not dados_dia:
                        continue
                    try:
                        if dados_dia.get("kg_eq_ref") not in (None, ""):
                            ws_me.cell(row=3, column=col, value=float(dados_dia["kg_eq_ref"]))
                            metricas_saurus += 1
                        if dados_dia.get("kg_eq_sob") not in (None, ""):
                            ws_me.cell(row=4, column=col, value=float(dados_dia["kg_eq_sob"]))
                            metricas_saurus += 1
                        if dados_dia.get("clientes") not in (None, ""):
                            ws_me.cell(row=5, column=col, value=int(dados_dia["clientes"]))
                            metricas_saurus += 1
                    except (ValueError, TypeError) as e:
                        logger.warning(f"[AVISO] Falha ao injetar métrica Saurus p/ {data_iso}: {e}")
                    alvos_verificacao.append({
                        "col_mensal": col,
                        "data_iso": data_iso,
                        "dados": dados_dia,
                    })
            if metricas_saurus:
                logger.info(
                    f"[+] Métricas Saurus (linhas 3/4/5) garantidas em {metricas_saurus} "
                    f"células para {self.aamm}."
                )
                self.stats["metricas_saurus"] = self.stats.get("metricas_saurus", 0) + metricas_saurus
            elif dados_cortex:
                logger.info(f"[+] Nenhuma métrica Saurus a injetar para {self.aamm}.")

            # ==============================================================================
            # ETAPA 2.6: DIVERGÊNCIA CAIXA x COMPUTADO (regra dos R$30)
            # ==============================================================================
            # Caixa (Movto_cx2, slips+dinheiro) já foi espelhado para a linha
            # LINHA_TOTAL_CAIXA do Diário. Computado (fechamento.txt/Saurus) vem
            # de dados_cortex[data]['total']. Compara os dois e grava/alerta.
            divergencias_flagged = 0
            if dados_cortex:
                mes_alvo = int(self.aamm[2:])
                ano_alvo = 2000 + int(self.aamm[:2])
                for dt, col in datas_no_mensal.items():
                    if dt.month != mes_alvo or dt.year != ano_alvo:
                        continue
                    if dt < DATA_MINIMA_PROCESSAMENTO:
                        continue
                    data_iso = dt.strftime("%Y-%m-%d")
                    dados_dia = dados_cortex.get(data_iso)
                    if not dados_dia:
                        continue  # dado indisponível na origem — não é divergência

                    valor_caixa = ws_me.cell(row=LINHA_TOTAL_CAIXA, column=col).value
                    try:
                        valor_caixa = float(valor_caixa) if valor_caixa not in (None, "") else None
                    except (ValueError, TypeError):
                        valor_caixa = None

                    valor_computado_raw = dados_dia.get("total")
                    try:
                        valor_computado = float(valor_computado_raw) if valor_computado_raw not in (None, "") else None
                    except (ValueError, TypeError):
                        valor_computado = None

                    categoria = registrar_divergencia(
                        data_iso, valor_caixa, valor_computado, bot=bot, chat_id=chat_id
                    )
                    registrar_checkup_dia(
                        data_iso, categoria,
                        detalhe=f"caixa={valor_caixa} computado={valor_computado}"
                    )
                    if categoria == "precisa_reconferencia":
                        divergencias_flagged += 1

                if divergencias_flagged:
                    logger.warning(
                        f"[divergencia] {divergencias_flagged} dia(s) do mês {self.aamm} "
                        f"acima de R${30:.2f} — reconferência manual necessária."
                    )
                self.stats["divergencias_flagged"] = divergencias_flagged

            # ==============================================================================
            # ETAPA 3: SALVAMENTO ESTRATÉGICO + VERIFICAÇÃO PÓS-ESCRITA (post-flight)
            # ==============================================================================
            try:
                wb_me.save(caminho_mensal)
                logger.info(f"\n[SUCESSO] Arquivo Movto_diario.{self.aamm}.xlsx salvo. Executando verificação pós-escrita...")
                self.status["save"] = "success"

                # Post-flight: reabre o arquivo e confirma que as células de métricas
                # (rows 3/4/5) foram de fato gravadas conforme o dados_cortex.
                self._verificar_pos_escrita(caminho_mensal, alvos_verificacao)

                return {
                    "status": "success",
                    "stats": self.stats,
                    "details": self.status
                }
            except PermissionError as e:
                # Arquivo provavelmente aberto no Excel: alerta claro antes de quebrar.
                alerta = (
                    f"[ALERTA CRÍTICO] Falha de escrita/verificação: Movto_diario.{self.aamm}.xlsx "
                    "provavelmente está aberto/travado no Excel. Feche-o e execute novamente."
                )
                logger.error(alerta)
                logger.error(f"Detalhe: {e}")
                self.status["save"] = "error: PermissionError (arquivo aberto no Excel)"
                raise RuntimeError(
                    "Falha crítica: O arquivo Movto_diario não foi preenchido corretamente."
                ) from e
            except RuntimeError as e:
                # Post-flight detectou divergência ou o PermissionError foi elevado.
                logger.error(f"[ALERTA CRÍTICO] {e}")
                self.status["save"] = "error"
                raise
            except Exception as e:
                error_msg = f"Erro ao salvar o arquivo Diário: {str(e)}"
                logger.error(error_msg)
                self.status["save"] = f"error: {str(e)}"
                raise RuntimeError(
                    "Falha crítica: O arquivo Movto_diario não foi preenchido corretamente."
                ) from e

        except Exception as e:
            error_msg = f"Erro inesperado no motor unificado: {str(e)}"
            logger.error(error_msg)
            return {"status": "error", "error": error_msg}

    def _verificar_pos_escrita(self, caminho_mensal: str, alvos: List[Dict[str, Any]]) -> None:
        """
        Post-flight check obrigatório: reabre o arquivo salvo e valida que as
        células de métricas Saurus (linhas 3/4/5) foram de fato gravadas.

        Lança RuntimeError (message conforme diretriz) se houver célula vazia,
        None ou valor divergente do `dados_cortex` original.

        Trata PermissionError de forma explícita: loga um alerta claro antes de
        quebrar a execução (arquivo provavelmente aberto no Excel).
        """
        if not alvos:
            return

        try:
            wb_check = load_workbook(caminho_mensal, data_only=False)
        except (PermissionError, FileNotFoundError, OSError) as e:
            # PermissionError = arquivo aberto/travado no Excel.
            # FileNotFoundError/OS = arquivo não encontrado ou travado por outro processo.
            if isinstance(e, PermissionError):
                logger.error(
                    f"[ALERTA CRÍTICO] Não foi possível reabrir {os.path.basename(caminho_mensal)} "
                    "para verificação pós-escrita (provavelmente aberto no Excel). "
                    "Feche o arquivo e tente novamente."
                )
            else:
                logger.error(
                    f"[ALERTA CRÍTICO] Falha ao reabrir {os.path.basename(caminho_mensal)} "
                    f"para verificação pós-escrita: {e}"
                )
            raise RuntimeError(
                "Falha crítica: O arquivo Movto_diario não foi preenchido corretamente."
            ) from e

        ws_check = wb_check.active

        # Mapa linha -> (campo do dados_dia, função de conversão)
        MAPA_ROWS = {
            3: ("kg_eq_ref", float),
            4: ("kg_eq_sob", float),
            5: ("clientes", int),
        }

        falhas: List[str] = []

        try:
            for alvo in alvos:
                col = alvo["col_mensal"]
                data_iso = alvo["data_iso"]
                dados = alvo["dados"]
                for row, (campo, conv) in MAPA_ROWS.items():
                    val = dados.get(campo)
                    if val in (None, ""):
                        continue
                    esperado = conv(val)
                    observado = ws_check.cell(row=row, column=col).value
                    if observado is None:
                        falhas.append(f"{data_iso} row {row}: célula vazia (esperado {esperado})")
                    else:
                        try:
                            obs_num = float(observado) if row != 5 else int(observado)
                        except (ValueError, TypeError):
                            falhas.append(f"{data_iso} row {row}: valor não numérico {observado!r}")
                            continue
                        if row != 5 and abs(obs_num - esperado) > 1e-6:
                            falhas.append(f"{data_iso} row {row}: esperado={esperado}, lido={obs_num}")
                        elif row == 5 and obs_num != esperado:
                            falhas.append(f"{data_iso} row {row}: esperado={esperado}, lido={obs_num}")
        finally:
            wb_check.close()

        if falhas:
            logger.error("[post-flight] Divergências detectadas: " + " | ".join(falhas))
            raise RuntimeError(
                "Falha crítica: O arquivo Movto_diario não foi preenchido corretamente."
            )
        logger.info("[post-flight] Verificação pós-escrita: todas as métricas confirmadas (rows 3/4/5).")

    def get_status(self) -> Dict[str, Any]:
        """Return the current status of the engine"""
        return {
            "status": self.status,
            "stats": self.stats,
            "aamm": self.aamm
        }

if __name__ == "__main__":
    logger.info("[*] ==================================================================")
    logger.info("[*] CÓRTEX PADROEIRA - MOTOR UNIFICADO (DIÁRIO MENSAL) - Async Version")
    logger.info("[*] ==================================================================")

    engine = EngineConsolidacaoAsync()
    result = engine.executar_motor_unificado()

    if result["status"] == "success":
        logger.info("Motor de consolidação executado com sucesso!")
        logger.info(f"Estatísticas: {engine.get_status()['stats']}")
    else:
        logger.error(f"Erro no motor de consolidação: {result['error']}")
