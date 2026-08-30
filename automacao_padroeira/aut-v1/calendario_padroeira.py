#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Calendário da Padroeira — regra de negócio (decision fcd26ad3, simplificada):
o restaurante NÃO abre em dias sem registro no caixa.

CRITÉRIO DEFINITIVO de "dia fechado" (decisão do Paulo em 10/08/2026):
AUSÊNCIA de registro no caixa — não existe fechamento_caixa_{data}.txt na
pasta fechamentos/ E não existe entrada no Movto_cx2.xlsx para aquela data.
Um calendário de feriados nacionais/municipais NÃO decide nada.

Domingo/feriado é usado apenas como SINAL AUXILIAR de log, para interpretar
a ausência (ex.: "dia X ausente, é domingo — provável fechado" vs
"dia X ausente, dia útil — possível bug").
"""

import functools
import os
from datetime import date, datetime, timedelta

# ---------------------------------------------------------------------------
# Feriados (NACIONAL fixo + MÓVEL) — usados APENAS como sinal auxiliar de log,
# nunca como critério de decisão de "dia fechado".
# ---------------------------------------------------------------------------
_FERIADOS_FIXOS = {
    (1, 1), (4, 21), (5, 1), (9, 7), (10, 12), (11, 2), (11, 15), (12, 25),
}

def _pascoa(ano: int) -> date:
    """Algoritmo de Gauss para a Páscoa (gregoriano)."""
    a = ano % 19
    b = ano // 100
    c = ano % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    mes = (h + l - 7 * m + 114) // 31
    dia = ((h + l - 7 * m + 114) % 31) + 1
    return date(ano, mes, dia)


def _feriados_moveis(ano: int) -> set:
    """Carnaval (terça), Sexta-feira Santa e Corpus Christi — só p/ log."""
    pascoa = _pascoa(ano)
    return {
        pascoa - timedelta(days=47),  # Carnaval (terça-feira)
        pascoa - timedelta(days=2),   # Sexta-feira Santa
        pascoa + timedelta(days=60),  # Corpus Christi
    }


def _mes_valido(aamm: str) -> bool:
    """Valida se o AAMM tem mês 01..12 (permite meses sintéticos 99 p/ testes)."""
    try:
        mes = int(aamm[2:])
        return 1 <= mes <= 12
    except ValueError:
        return False


def _data_str(aamm: str, dia: int) -> str:
    """Converte AAMM+dia em 'AAAA-MM-DD' (ex.: 2607,5 -> '2026-07-05')."""
    return f"{2000 + int(aamm[:2]):04d}-{int(aamm[2:]):02d}-{dia:02d}"


# ---------------------------------------------------------------------------
# CRITÉRIO DEFINITIVO: presença/ausência de registro no caixa
# ---------------------------------------------------------------------------
def tem_fechamento_caixa(base_dir: str, aamm: str, dia: int) -> bool:
    """True se existe fechamento_caixa_{data}.txt na pasta fechamentos/."""
    caminho = os.path.join(
        base_dir, "fechamentos", f"fechamento_caixa_{_data_str(aamm, dia)}.txt"
    )
    return os.path.exists(caminho)


# ---------------------------------------------------------------------------
# Cache LRU do CABEÇALHO de datas do Movto_cx2.xlsx (Item 2 — refatorar.md).
#
# ANTES: cada dia validado do mês reabria o .xlsx do disco (load_workbook em
# read_only + iteração da linha 1) -> dezenas de aberturas por varredura de mês.
#
# AGORA: o cabeçalho (conjunto de datas presente na linha 1) é carregado UMA
# vez por arquivo (keyed por mtime+sig) e缓存 em LRU. `tem_movimento_cx2` passa
# a apenas consultar o conjunto em memória. Quem já tem o workbook carregado
# pode injetá-lo via `workbook=` (evita até o load).
# ---------------------------------------------------------------------------
@functools.lru_cache(maxsize=8)
def _cabecalho_cx2(caminho: str, _mtime: float) -> frozenset:
    """
    Retorna o conjunto (imutável) de `date` presentes na linha 1 do Movto_cx2.
    O `_mtime` é parte da chave do cache para invalidar quando o arquivo muda.
    Em caso de erro de leitura, retorna conjunto vazio (seguro: 'não tem
    movimento').
    """
    datas: set = set()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(caminho, data_only=True, read_only=True)
        try:
            ws = wb.active
            for row in ws.iter_rows(min_row=1, max_row=1, min_col=2):
                for cel in row:
                    v = cel.value
                    # openpyxl devolve datetime.datetime para células de data;
                    # normalizamos para date para a busca em memória ser consistente.
                    if isinstance(v, datetime):
                        datas.add(v.date())
        finally:
            wb.close()
    except Exception:
        pass
    return frozenset(datas)


def _caminho_cx2(base_dir: str) -> str:
    return os.path.join(base_dir, "Movto_cx2.xlsx")


def tem_movimento_cx2(base_dir: str, aamm: str, dia: int, *, workbook=None) -> bool:
    """
    True se o Movto_cx2.xlsx tem uma coluna com a data (ano/mês/dia).

    ITEM 2 (refatorar.md): aceita `workbook` já carregado (injeção, evita
    reabrir o disco) OU usa cache LRU do cabeçalho (não reabre por dia).
    """
    if not _mes_valido(aamm):
        return False
    ano = 2000 + int(aamm[:2])
    mes = int(aamm[2:])

    # Modo injeção: workbook já carregado pelo chamador.
    if workbook is not None:
        try:
            ws = workbook.active
            for row in ws.iter_rows(min_row=1, max_row=1, min_col=2):
                for cel in row:
                    v = cel.value
                    if isinstance(v, datetime):
                        v = v.date()
                    if isinstance(v, date) and v.year == ano and v.month == mes and v.day == dia:
                        return True
        except Exception:
            pass
        return False

    # Modo cache LRU: carrega/aproveita o cabeçalho e consulta em memória.
    caminho = _caminho_cx2(base_dir)
    if not os.path.exists(caminho):
        return False
    try:
        mtime = os.path.getmtime(caminho)
    except OSError:
        return False
    datas = _cabecalho_cx2(caminho, mtime)
    try:
        return date(ano, mes, dia) in datas
    except ValueError:
        # Dia impossível para o mês (ex.: 30/02) — com certeza não há movimento.
        return False


def dia_eh_fechado(base_dir: str, aamm: str, dia: int) -> bool:
    """
    CRITÉRIO DEFINITIVO: dia fechado = AUSÊNCIA de registro no caixa.
    Fechado se NÃO há fechamento_caixa_{data}.txt E NÃO há entrada no
    Movto_cx2.xlsx para a data. O calendário/feriado não decide.
    """
    if not _mes_valido(aamm):
        return False
    return not (tem_fechamento_caixa(base_dir, aamm, dia)
                or tem_movimento_cx2(base_dir, aamm, dia))


def sinal_auxiliar(aamm: str, dia: int) -> str:
    """
    Sinal auxiliar de LOG (não decide): 'domingo', 'feriado' ou 'dia útil'.
    Ex.: 'domingo — provável fechado' vs 'dia útil — possível bug'.
    """
    if not _mes_valido(aamm):
        return "dia útil"
    ano = 2000 + int(aamm[:2])
    mes = int(aamm[2:])
    try:
        data = date(ano, mes, dia)
    except ValueError:
        return "dia útil"
    if data.weekday() == 6:
        return "domingo"
    if data in _feriados_fixos(ano):
        return "feriado"
    if data in _feriados_moveis(ano):
        return "feriado"
    return "dia útil"


def _feriados_fixos(ano: int) -> set:
    return {date(ano, mes, dia) for mes, dia in _FERIADOS_FIXOS}


def dias_do_mes(aamm: str) -> int:
    """Quantidade de dias do mês do AAMM (meses sintéticos de teste → 31)."""
    if not _mes_valido(aamm):
        return 31
    import calendar
    ano = 2000 + int(aamm[:2])
    mes = int(aamm[2:])
    return calendar.monthrange(ano, mes)[1]
