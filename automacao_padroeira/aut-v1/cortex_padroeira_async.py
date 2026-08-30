#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Córtex Principal - Automação Ecossistema Padroeira (Async Version)
Orquestrador Central adaptado para Async Reconciliation Architecture V2

Versão reestruturada para extração multi-data:
  - Para cada data pendente, tenta carregar fechamentos/fechamento_caixa_{AAA_MM_DD}.txt.
  - Se ausente, invoca o robô Playwright (pdv_saurus_extractor) para resgatar
    o relatório do portal Saurus e salvá-lo no cache local.
  - Fallback para fechamento_caixa.txt estático quando Playwright não disponível
    ou sem credenciais no .env.
"""

import asyncio
import glob
import os
import re
from datetime import datetime
from typing import Dict, Any, List, Optional

import telebot
from openpyxl import load_workbook
import logging

from calendario_padroeira import dia_eh_fechado

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger("CortexPadroeiraAsync")

# Caminhos dinâmicos relativos ao próprio script (portável, sem hardcode).
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Data mínima de processamento: nada anterior a isso deve ser tocado pelo
# pipeline, mesmo em uma reexecução "do zero" — meses anteriores já estão
# fechados/conferidos manualmente e ficam fora de escopo.
DATA_MINIMA_PROCESSAMENTO = datetime(2026, 6, 1).date()


def _ler_env(chave: str, base_dir: str) -> Optional[str]:
    """Lê `chave` do ambiente, ou do arquivo .env local como fallback."""
    val = os.environ.get(chave)
    if val is not None:
        return val
    env_path = os.path.join(base_dir, ".env")
    if os.path.exists(env_path):
        try:
            with open(env_path, "r", encoding="utf-8") as f:
                for linha in f:
                    k, _, v = linha.partition("=")
                    if k.strip().upper() == chave.upper():
                        return v.strip().strip('"').strip("'")
        except Exception as e:
            logger.warning(f"[AVISO] Falha ao ler .env para {chave}: {e}")
    return None


# CONFIGURAÇÃO DE AMBIENTE
# O token NUNCA fica no código — vem de TELEGRAM_TOKEN no ambiente ou no .env
# local (fora do git). Se o token antigo (que estava hardcoded aqui) ainda não
# foi revogado no BotFather, revogue-o agora: ele foi exposto no código-fonte.
TOKEN_TELEGRAM = _ler_env("TELEGRAM_TOKEN", BASE_DIR)
CHAT_ID_ALERTAS = _ler_env("TELEGRAM_CHAT_ID", BASE_DIR)  # chat/grupo que recebe alertas de divergência

if not TOKEN_TELEGRAM:
    raise RuntimeError(
        "TELEGRAM_TOKEN não encontrado no ambiente nem no .env. "
        "Configure-o antes de iniciar o bot (nunca deixe o token hardcoded no código)."
    )

bot = None  # instanciado sob demanda em main(); evita objeto duplicado com
             # bot_reconciliation.py (que cria o seu próprio bot no mesmo token).


def _iterar_cabecalho(ws, inicio: int = 1):
    """
    Itera as células da linha 1 a partir de `inicio`, aplicando parada rápida:
    ao encontrar 2 colunas consecutivas vazias/None, interrompe a varredura.
    Evita iterar as ~16k colunas vazias finais do Movto_cx2.xlsx.
    """
    vazios = 0
    for c in range(inicio, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v in (None, ""):
            vazios += 1
            if vazios >= 2:
                break
            continue
        vazios = 0
        yield c, v


class CortexPadroeiraAsync:
    """
    Async version of Cortex Padroeira for integration with Async Reconciliation V2.
    Suporta extração multi-data com cache em pasta ./fechamentos/.
    """

    def __init__(self, base_dir: Optional[str] = None):
        self.base_dir = base_dir or BASE_DIR
        self.pasta_fechamentos = os.path.join(self.base_dir, "fechamentos")
        os.makedirs(self.pasta_fechamentos, exist_ok=True)

        self.status: Dict[str, Any] = {"data_extraction": None, "planilha_check": None}
        self.dados: Optional[Dict[str, str]] = None          # compat: último extraído (ou fallback)
        self.dados_por_data: Dict[str, Dict[str, str]] = {}   # dados por data (AAAA-MM-DD)
        self.pendentes: List[str] = []                         # datas pendentes

    # ------------------------------------------------------------------
    # Extração por data (sync, parse de arquivo)
    # ------------------------------------------------------------------
    def extrair_dados_saurus_por_data(self, data_str: str) -> Optional[Dict[str, str]]:
        """
        Lê o arquivo fechamento_caixa_{data_str}.txt ou, como fallback,
        o arquivo padrão fechamento_caixa.txt; parseia e retorna o
        dicionário de totais para aquela data específica.
        """
        txt_path = os.path.join(self.pasta_fechamentos, f"fechamento_caixa_{data_str}.txt")

        # Fallback: arquivo padrão (único arquivo estático legado)
        fallback_path = os.path.join(self.base_dir, "fechamento_caixa.txt")
        if not os.path.exists(txt_path):
            if os.path.exists(fallback_path):
                logger.warning(
                    f"[AVISO] Usando fechamento_caixa.txt padrão para a data {data_str}"
                )
                txt_path = fallback_path
            else:
                logger.warning(f"Nenhum arquivo de fechamento encontrado para {data_str}")
                return None

        try:
            with open(txt_path, "r", encoding="utf-8") as f:
                conteudo = f.read()
            return self._parsear_fechamento(conteudo, data_str)
        except Exception as e:
            logger.error(f"Erro ao extrair dados para {data_str}: {e}")
            return None

    def _parsear_fechamento(self, conteudo: str, data_str: str) -> Optional[Dict[str, str]]:
        """Aplica as regex do formato Saurus no conteúdo bruto do arquivo.

        Além dos pesos brutos (linhas do BALANÇA), extrai as quantidades
        unitárias (A Vontade, To Save), os valores em R$ das categorias
        PRATOS EXECUTIVOS e DOCES, e calcula o "Kg Equivalente" das
        refeições (linha 3) e das sobremesas/doces (linha 4).
        """
        # Formato real dos relatórios Saurus: "DINHEIRO (14):  595.76" — o número
        # entre parênteses é o contador de vendas daquela forma. Tornamos os
        # parênteses + contador OPCIONAIS (podem não vir) e capturamos o VALOR após
        # os dois pontos. Antes, a regex exigia "DINHEIRO 14:" sem parênteses e,
        # como o relatório traz "(14)", o match quebrava e os valores financeiros
        # (Total/Dinheiro/Crédito/Débito) caíam no default "0.00" — bug que afetava
        # a injeção no Diário/PAD. (corrigido na sessão 2026-08-27)
        #
        # CORREÇÃO P1 (refatorar.md 2b): dinheiro/credito/debito usavam classe
        # [\d.]+ (sem vírgula). Se o relatório vier em formato pt-BR com vírgula
        # decimal (ex.: "DINHEIRO (14): 1.234,56"), capturava apenas "1.234" e
        # quebrava a conversão — corrompendo justamente a comparação de DINHEIRO
        # da ETAPA 2.6. Alinhamos à mesma classe [\d.,]+ do `total` e normalizamos
        # com .replace(",", ".") antes do float.
        _RE_VALOR = r"([\d.,]+)"

        def _valor_financeiro(label: str) -> Optional[str]:
            m = re.search(rf"{label}(?:\s+\(\d+\))?\s*:\s*{_RE_VALOR}", conteudo)
            if not m:
                return None
            return m.group(1).replace(",", ".")

        dinheiro = _valor_financeiro("DINHEIRO")
        credito  = _valor_financeiro("CRÉDITO")
        debito   = _valor_financeiro("DÉBITO")
        total    = _valor_financeiro("TOTAL")
        clientes = re.search(r"Qtd\. Vendas\s+:\s+(\d+)", conteudo)
        # IMPORTANTE: podem existir VARIAS linhas "REFEICAO QUILO KG" / "SOBREMESA QUILO KG"
        # no mesmo fechamento (ex.: almoco + jantar). Por isso usamos findall + soma,
        # e nao re.search (que pegaria so a primeira e subestimaria o peso do dia).
        def _somar_kg(padrao: str) -> float:
            vals = re.findall(padrao, conteudo)
            soma = 0.0
            for v in vals:
                try:
                    soma += float(v.replace(",", "."))
                except ValueError:
                    continue
            return soma

        def _num(padrao: str, flags: int = 0) -> Optional[float]:
            m = re.search(padrao, conteudo, flags)
            if not m:
                return None
            try:
                return float(m.group(1).replace(",", "."))
            except ValueError:
                return None

        peso_buf = _somar_kg(r"REFEICAO QUILO\s+KG\s+([\d.,]+)")
        peso_sob = _somar_kg(r"SOBREMESA QUILO\s+KG\s+([\d.,]+)")

        # --- Quantidades unitárias e valores em R$ para o Kg Equivalente ---
        qtd_av = _num(r"REFEICAO A VONTADE\s+UN\s+([\d.,]+)")
        qtd_ts = _num(r"REFEICAO TO SAVE\s+UN\s+([\d.,]+)")
        qtd_cs = _num(r"REFEICAO COM SOBREMESA\s+UN\s+([\d.,]+)")
        # PRATOS EXECUTIVOS e DOCES: valor em R$ na seção "SUBCATEGORIAS VENDIDAS".
        r_exec = re.search(r"PRATOS EXECUTIVOS\s+[\d.,]+\s+([\d.,]+)", conteudo)
        r_doces = re.search(r"\bDOCES\s+[\d.,]+\s+([\d.,]+)", conteudo)
        val_exec = float(r_exec.group(1).replace(",", ".")) if r_exec else 0.0
        val_doces = float(r_doces.group(1).replace(",", ".")) if r_doces else 0.0

        # Preço do KG do dia (regra de dia da semana / override por data em config_precos).
        try:
            from config_precos import valor_kg_dia, REFEICAO_COM_SOBREMESA
            dt_dia = datetime.strptime(data_str, "%Y-%m-%d").date()
            vkg = valor_kg_dia(dt_dia)
        except Exception:
            vkg = 96.90  # fallback conservador (padrão dias úteis)
            REFEICAO_COM_SOBREMESA = 73.90

        # --- Kg Equivalente: Refeição (linha 3) ---
        # Inclui REFEIÇÃO COM SOBREMESA (un), que antes era ignorada — valor fixo
        # unitário desse item soma ao equivalente do dia. (rega: vkrisma 03 e 07/08)
        fat_ref = (peso_buf * vkg) + (qtd_av or 0.0) * 63.90 + (qtd_ts or 0.0) * 13.90 \
            + (qtd_cs or 0.0) * float(REFEICAO_COM_SOBREMESA) + val_exec
        kg_eq_ref = fat_ref / vkg if vkg else 0.0
        # --- Kg Equivalente: Sobremesa / Doces (linha 4) ---
        fat_sob = (peso_sob * vkg) + val_doces
        kg_eq_sob = fat_sob / vkg if vkg else 0.0

        dados = {
            "data": data_str,
            "dinheiro": dinheiro if dinheiro else "0.00",
            "credito":  credito  if credito  else "0.00",
            "debito":   debito   if debito   else "0.00",
            "total":    total if total else "0.00",
            "total_bruto": total if total else "0.00",
            "clientes": clientes.group(1) if clientes else "0",
            "peso_buf": f"{peso_buf:.3f}",
            "peso_sob": f"{peso_sob:.3f}",
            # Kg Equivalente (convertido R$ -> kg unificado)
            "kg_eq_ref": f"{kg_eq_ref:.3f}",
            "kg_eq_sob": f"{kg_eq_sob:.3f}",
            # Detalhe para auditoria/debug (nao injetado na planilha)
            "_kg_eq_debug": {
                "vkg": round(vkg, 2),
                "qtd_av": qtd_av, "qtd_ts": qtd_ts,
                "val_exec": round(val_exec, 2), "val_doces": round(val_doces, 2),
                "fat_ref": round(fat_ref, 2), "fat_sob": round(fat_sob, 2),
            },
        }
        self.dados_por_data[data_str] = dados
        return dados

    # ------------------------------------------------------------------
    # Extração legada (arquivo único) — mantida por compatibilidade
    # ------------------------------------------------------------------
    def extrair_dados_saurus(self) -> Optional[Dict[str, str]]:
        """Lê fechamento_caixa.txt e retorna totais. Usado pelo /fechar (Telegram)."""
        dados = self.extrair_dados_saurus_por_data("_legado_")
        if dados:
            self.dados = dados
            self.status["data_extraction"] = "success"
            return dados
        self.status["data_extraction"] = "error: fechamento_caixa.txt não encontrado"
        return None

    # ------------------------------------------------------------------
    # Cache local de fechamentos (fonte para as linhas 3/4/5)
    # ------------------------------------------------------------------
    def carregar_cache_fechamentos(self) -> int:
        """
        Varre a pasta ./fechamentos/ e carrega TODOS os fechamentos locais
        (fechamento_caixa_*.txt) em self.dados_por_data.

        Não depende das datas "pendentes" do calendário: garante que o cache
        local esteja disponível para a injeção das linhas 3/4/5 mesmo quando
        todas as colunas do Diário já existem (porém vazias).
        """
        padrao = os.path.join(self.pasta_fechamentos, "fechamento_caixa_*.txt")
        carregados = 0
        for arquivo in sorted(glob.glob(padrao)):
            data_str = os.path.basename(arquivo).replace("fechamento_caixa_", "").replace(".txt", "")
            if data_str == "_legado_":
                continue
            try:
                with open(arquivo, "r", encoding="utf-8") as f:
                    conteudo = f.read()
                if self._parsear_fechamento(conteudo, data_str):
                    carregados += 1
            except Exception as e:
                logger.error(f"[cache] Falha ao carregar fechamento {data_str}: {e}")
        if carregados:
            logger.info(f"[cache] {carregados} fechamento(s) carregado(s) do cache local.")
        else:
            logger.info("[cache] Nenhum fechamento local encontrado em ./fechamentos/.")

        return carregados

    # ------------------------------------------------------------------
    # Verificação de paridade
    # ------------------------------------------------------------------
    def verificar_paridade_planilhas(self, aamm: Optional[str] = None) -> bool:
        """Verifica a paridade entre Movto_cx2.xlsx e Movto_diario.{AAMM}.xlsx."""
        try:
            cx2_path    = os.path.join(self.base_dir, "Movto_cx2.xlsx")
            aamm_alvo   = aamm or datetime.now().strftime("%y%m")
            diario_path = os.path.join(self.base_dir, f"Movto_diario.{aamm_alvo}.xlsx")

            if not os.path.exists(diario_path):
                logger.error(f"Planilha '{os.path.basename(diario_path)}' não encontrada")
                self.status["planilha_check"] = f"error: {os.path.basename(diario_path)} not found"
                return False

            wb_cx2 = load_workbook(cx2_path, data_only=True)
            ws_cx2 = wb_cx2.active
            datas_cx2 = [v for _, v in _iterar_cabecalho(ws_cx2, inicio=2) if v]

            wb_diario = load_workbook(diario_path, data_only=True)
            ws_diario = wb_diario.active
            datas_diario = [v for _, v in _iterar_cabecalho(ws_diario, inicio=2) if v]

            str_cx2    = {d.strftime("%Y-%m-%d") for d in datas_cx2    if isinstance(d, datetime)}
            str_diario = {d.strftime("%Y-%m-%d") for d in datas_diario if isinstance(d, datetime)}

            # Filtra para o período AAMM alvo: o Movto_cx2 concentra datas de VÁRIOS
            # meses; só interessam como pendentes as datas do período em análise.
            str_cx2 = {
                d for d in str_cx2
                if datetime.strptime(d, "%Y-%m-%d").strftime("%y%m") == aamm_alvo
            }

            # Regra de negócio (fcd26ad3 simplificada): "dia fechado" = AUSÊNCIA de
            # registro no caixa (sem fechamento_caixa_{data}.txt E sem entrada no
            # Movto_cx2). Datas presentes no Movto_cx2 têm registro no caixa, logo
            # são pendentes legítimas; domingo/feriado não decide mais.
            pendentes_bruto = sorted(list(str_cx2 - str_diario))
            fechados = []
            pendentes = []
            fora_de_escopo = []
            for dt in pendentes_bruto:
                data_dt = datetime.strptime(dt, "%Y-%m-%d").date()
                if data_dt < DATA_MINIMA_PROCESSAMENTO:
                    # Fora de escopo: meses anteriores a jun/2026 não são
                    # tocados, mesmo que apareçam como "pendentes" no Caixa 2.
                    fora_de_escopo.append(dt)
                    continue
                dia = datetime.strptime(dt, "%Y-%m-%d").day
                # Como `dt` veio do próprio Movto_cx2, tem registro no caixa -> pendente.
                if dia_eh_fechado(self.base_dir, aamm_alvo, dia):
                    fechados.append(dt)
                else:
                    pendentes.append(dt)

            if fora_de_escopo:
                logger.info(
                    f"[escopo] {len(fora_de_escopo)} data(s) anteriores a "
                    f"{DATA_MINIMA_PROCESSAMENTO.isoformat()} ignoradas (fora de escopo): "
                    f"{sorted(fora_de_escopo)}"
                )

            if fechados:
                logger.info(
                    f"[paridade] {len(fechados)} data(s) tratadas como fechado "
                    f"(sem registro no caixa): {sorted(fechados)}"
                )
            self.pendentes = pendentes
            self.status["planilha_check"] = "success"
            return True

        except Exception as e:
            logger.error(f"Erro ao verificar paridade de planilhas: {e}")
            self.status["planilha_check"] = f"error: {e}"
            return False

    # ------------------------------------------------------------------
    # Detecção de pendências de EXTRAÇÃO (coluna existe, falta Saurus)
    # ------------------------------------------------------------------
    def _datas_sem_metricas_saurus(self, aamm_alvo: str) -> List[str]:
        """
        Datas cuja coluna JÁ existe no Movto_diario.{aamm} (o caixa foi
        espelhado pelo Engine), MAS sem métricas do Saurus (linhas 3/4/5
        vazias) e sem fechamento_caixa_{data}.txt em cache.

        São pendentes de EXTRAÇÃO via Playwright: a paridade por presença de
        coluna (cx2 - diario) não as pega, pois a coluna já existe. Sem esta
        detecção, o mês "parece" 100% consolidado quando só o caixa entrou.
        """
        diario_path = os.path.join(self.base_dir, f"Movto_diario.{aamm_alvo}.xlsx")
        if not os.path.exists(diario_path):
            return []
        wb = load_workbook(diario_path, data_only=True)
        ws = wb.active
        out = []
        for _c, v in _iterar_cabecalho(ws, inicio=2):
            if not isinstance(v, datetime):
                continue
            data_iso = v.strftime("%Y-%m-%d")
            try:
                if datetime.strptime(data_iso, "%Y-%m-%d").strftime("%y%m") != aamm_alvo:
                    continue
                if datetime.strptime(data_iso, "%Y-%m-%d").date() < DATA_MINIMA_PROCESSAMENTO:
                    continue
            except ValueError:
                continue
            # Já tem .txt no cache? Não é pendente de extração.
            if os.path.exists(
                os.path.join(self.pasta_fechamentos, f"fechamento_caixa_{data_iso}.txt")
            ):
                continue
            # Linhas 3/4/5 preenchidas? Saurus já foi injetado.
            col = _c
            if any(ws.cell(row=ln, column=col).value not in (None, "") for ln in (3, 4, 5)):
                continue
            out.append(data_iso)
        wb.close()
        return out

    # ------------------------------------------------------------------
    # Extração multi-data (async, com Playwright)
    # ------------------------------------------------------------------
    def _playwright_disponivel(self) -> bool:
        """Verifica se Playwright está instalado e acessível."""
        try:
            from playwright.async_api import async_playwright  # noqa: F401
            return True
        except ImportError:
            return False

    def _headless_config(self) -> bool:
        """
        Resolve o modo headless do Playwright a partir de PLAYWRIGHT_HEADLESS,
        consultando primeiro o ambiente e depois o .env local.
        Default True (produção / servidor sem X server).
        """
        val = os.environ.get("PLAYWRIGHT_HEADLESS")
        if val is None:
            env_path = os.path.join(self.base_dir, ".env")
            if os.path.exists(env_path):
                try:
                    with open(env_path, "r", encoding="utf-8") as f:
                        for linha in f:
                            chave, _, v = linha.partition("=")
                            if chave.strip().upper() == "PLAYWRIGHT_HEADLESS":
                                val = v.strip().strip('"').strip("'")
                                break
                except Exception as e:
                    logger.warning(f"[AVISO] Falha ao ler .env para headless: {e}")
        if val is None:
            return True
        return val.strip().lower() in ("1", "true", "yes", "on")

    async def extrair_todos_pendentes(self, headless: bool = False, limite: Optional[int] = None) -> Dict[str, Optional[str]]:
        """
        Para cada data em self.pendentes, extrai (ou lê cache) e armazena
        em self.dados_por_data. Retorna {data_str: caminho_arquivo ou None}.

        As datas SEM cache são extraídas em UMA ÚNICA sessão de browser
        (reaproveitando login — ver `extrator_saurus_sessao.extrair_lote_saurus`,
        a versão comprovada que produziu 264/0 relatórios em 08/ago/2026).
        Isso evita abrir/fechar o browser por data e disparar alarme de
        segurança do portal.

        Args:
            headless: roda sem display (padrão True em servidor).
            limite: se informado, restringe a amostra de datas a extrair
                    (útil p/ testes e2e de 2-3 dias).
        """
        resultados: Dict[str, Optional[str]] = {}

        # 1) Cache: datas que já têm .txt no disco
        pendentes_extracao = []
        for dt in self.pendentes:
            cache = os.path.join(self.pasta_fechamentos, f"fechamento_caixa_{dt}.txt")
            if os.path.exists(cache):
                logger.info(f"[cache] Fechamento já existe para {dt}: {os.path.basename(cache)}")
                self.extrair_dados_saurus_por_data(dt)
                resultados[dt] = cache
            else:
                pendentes_extracao.append(dt)

        if limite is not None and len(pendentes_extracao) > limite:
            pendentes_extracao = pendentes_extracao[:limite]

        # 2) Sessão única do Playwright para as datas pendentes
        if pendentes_extracao and self._playwright_disponivel():
            from extrator_saurus_sessao import extrair_lote_saurus
            try:
                ok, falhas = await extrair_lote_saurus(
                    pendentes_extracao,
                    self.pasta_fechamentos,
                    headless=headless,
                    on_progress=lambda i, tot, d, okp: logger.info(
                        f"[PLAYWRIGHT] Baixando fechamento para a data "
                        f"{self._fmt_data_br(d)} ({i}/{tot}) -> {'OK' if okp else 'FALHA'}"
                    ),
                )
                logger.info(f"[PLAYWRIGHT] Lote concluído: {ok} OK, {falhas} falhas "
                            f"de {len(pendentes_extracao)} datas.")
            except Exception as e:
                logger.error(f"[PLAYWRIGHT] Falha no lote do portal Saurus: {e}")

        # 3) Pós-extração: parseia cada data (cache ou recém-baixada)
        for dt in pendentes_extracao:
            cache = os.path.join(self.pasta_fechamentos, f"fechamento_caixa_{dt}.txt")
            if os.path.exists(cache):
                self.extrair_dados_saurus_por_data(dt)
                resultados[dt] = cache
            else:
                dados = self.extrair_dados_saurus_por_data(dt)
                resultados[dt] = None if dados is None else cache

        return resultados

    @staticmethod
    def _fmt_data_br(data_iso: str) -> str:
        """Converte '2026-08-04' -> '04/08/2026' para logs amigáveis."""
        try:
            from datetime import datetime as _dt
            return _dt.strptime(data_iso, "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            return data_iso

    # ------------------------------------------------------------------
    # Status
    # ------------------------------------------------------------------
    def get_status(self) -> Dict[str, Any]:
        return {
            "status": self.status,
            "dados": self.dados,
            "dados_por_data": self.dados_por_data,
            "pendentes": self.pendentes,
            "has_pending_dates": len(self.pendentes) > 0,
        }


# ----------------------------------------------------------------------
# Telegram Bot Handlers (uso standalone com /fechar e /ok)
#
# IMPORTANTE (P1 - refatorar.md): o bot NÃO é instanciado no nível de módulo.
# Antes, `bot = telebot.TeleBot(...)` era criado globalmente, e ao ser importado
# por bot_reconciliation.py gerava um SEGUNDO objeto bot com o mesmo token
# (além do que o próprio bot_reconciliation cria), duplicando handlers "mortos"
# e conflitando com a escuta principal. Agora a instância e os handlers só
# existem dentro de main(), que roda exclusivamente em __main__.
# ----------------------------------------------------------------------
def main() -> None:
    global bot
    bot = telebot.TeleBot(TOKEN_TELEGRAM)
    cortex_async = CortexPadroeiraAsync()

    @bot.message_handler(commands=['fechar'])
    def comando_fechar(message):
        bot.reply_to(message, "🤖 *Córtex Lab:* Processando dados brutos do Saurus...")
        dados = cortex_async.extrair_dados_saurus()

        if not dados:
            bot.reply_to(message, "❌ Erro: Arquivo 'fechamento_caixa.txt' não encontrado na pasta do laboratório.")
            return

        msg = (
            f"📊 *FATURAMENTO TOTAL DO DIA*\n\n"
            f"• DINHEIRO: R$ {dados['dinheiro']}\n"
            f"• CRÉDITO: R$ {dados['credito']}\n"
            f"• DÉBITO: R$ {dados['debito']}\n"
            f"• *TOTAL DO SISTEMA: R$ {dados['total']}*\n\n"
            f"⚖️ *Métricas Equivalentes (Automação):*\n"
            f"• Refeição Quilo: {dados['peso_buf']} KG\n"
            f"• Sobremesa Quilo: {dados['peso_sob']} KG\n"
            f"• Número de Clientes: {dados['clientes']}\n\n"
            f"✍️ Sandra, preencha o *Movto_cx2.xlsx* no PC.\n"
            f"Quando terminar e salvar, digite */ok* aqui para consolidar!"
        )
        bot.send_message(message.chat.id, msg, parse_mode="Markdown")

    @bot.message_handler(commands=['ok'])
    def comando_ok(message):
        bot.reply_to(message, "⚡ *Córtex Lab:* Comparando planilhas e checando paridade...")

        if not cortex_async.verificar_paridade_planilhas():
            bot.reply_to(message, "❌ Erro ao verificar paridade de planilhas.")
            return

        if cortex_async.pendentes:
            bot.send_message(message.chat.id, f"🔄 *Datas pendentes detectadas no Caixa 2:* {len(cortex_async.pendentes)} dia(s).")
        else:
            bot.send_message(message.chat.id, "✅ Paridade total encontrada! Nenhuma data nova detectada.")

        bot.send_message(message.chat.id, "🚀 *Processo de reconciliação assíncrona iniciado!* Aguarde a conclusão...")

    logger.info("[*] LAB CÓRTEX (Async): Ativo e escutando o Telegram...")
    bot.infinity_polling()


if __name__ == "__main__":
    main()
