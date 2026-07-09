#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Córtex Principal - Automação Ecossistema Padroeira
Versão síncrona para integração com a Arquitetura de Reconciliação V2
"""

import os
import re
from datetime import datetime
import telebot
from openpyxl import load_workbook
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger("CortexPadroeira")

# CONFIGURAÇÃO DE AMBIENTE
TOKEN_TELEGRAM = "8890227531:AAEWbuEyzA3G2VImmYICJBK8ZqCUsBbq9Ro"  # ← SUBSTITUA PELO SEU TOKEN REAL

# AJUSTE PARA MOCK_BOX DURANTE DESENVOLVIMENTO
PASTA_LAB = os.path.expanduser("~/work_out")
PASTA_VENDAS_ORIGEM = os.path.join(PASTA_LAB, "mock_box", "Padroeira_vendas")
PASTA_RESTAURANTE_ANO = os.path.join(PASTA_LAB, "mock_box", "Restaurante", "A2026")

bot = telebot.TeleBot(TOKEN_TELEGRAM)

class CortexPadroeira:
    """
    Córtex Padroeira para integração com a Arquitetura de Reconciliação V2
    """

    def __init__(self):
        self.status = {"data_extraction": None, "planilha_check": None}
        self.dados = None
        self.pendentes = []

    def extrair_dados_saurus(self) -> dict:
        """Lê o fechamento_caixa.txt e calcula os totais e equivalentes de peso"""
        txt_path = os.path.join(PASTA_LAB, "fechamento_caixa.txt")
        if not os.path.exists(txt_path):
            logger.error(f"Arquivo 'fechamento_caixa.txt' não encontrado em {txt_path}")
            return None

        try:
            with open(txt_path, "r", encoding="utf-8") as f:
                conteudo = f.read()

            # Busca valores financeiros usando Regex simples
            dinheiro = re.search(r"DINHEIRO:\s+R\$\s+([\d\.,]+)", conteudo)
            credito = re.search(r"CR[ÉE]DITO:\s+R\$\s+([\d\.,]+)", conteudo)
            debito = re.search(r"D[ÉE]BITO:\s+R\$\s+([\d\.,]+)", conteudo)
            total = re.search(r"TOTAL:\s+R\$\s+([\d\.,]+)", conteudo)

            # Busca Pesos e Clientes
            clientes = re.search(r"Total de clientes:\s+(\d+)", conteudo)
            peso_buf = re.search(r"Buf:\s+R\$\s+([\d\.,]+)", conteudo)
            peso_sob = re.search(r"Sob:\s+R\$\s+([\d\.,]+)", conteudo)

            self.dados = {
                "dinheiro": dinheiro.group(1) if dinheiro else "0.00",
                "credito": credito.group(1) if credito else "0.00",
                "debito": debito.group(1) if debito else "0.00",
                "total": total.group(1).replace(",", ".") if total else "0.00",
                "clientes": clientes.group(1) if clientes else "0",
                "peso_buf": peso_buf.group(1).replace(",", ".") if peso_buf else "0.000",
                "peso_sob": peso_sob.group(1).replace(",", ".") if peso_sob else "0.000"
            }

            self.status["data_extraction"] = "success"
            return self.dados

        except Exception as e:
            logger.error(f"Erro ao extrair dados do Saurus: {str(e)}")
            self.status["data_extraction"] = f"error: {str(e)}"
            return None

    def verificar_paridade_planilhas(self) -> bool:
        """Verifica a paridade entre as planilhas Movto_cx2.xlsx e Movto_diario"""
        try:
            cx2_path = os.path.join(PASTA_VENDAS_ORIGEM, "Movto_cx2.xlsx")
            aamm_atual = datetime.now().strftime("%y%m")
            diario_path = os.path.join(PASTA_RESTAURANTE_ANO, f"Movto_diario.{aamm_atual}.xlsx")

            if not os.path.exists(diario_path):
                logger.error(f"Planilha de teste '{os.path.basename(diario_path)}' não encontrada no mock_box")
                self.status["planilha_check"] = f"error: {os.path.basename(diario_path)} not found"
                return False

            wb_cx2 = load_workbook(cx2_path, data_only=True)
            ws_cx2 = wb_cx2.active
            datas_cx2 = [ws_cx2.cell(row=1, column=c).value for c in range(2, ws_cx2.max_column + 1) if ws_cx2.cell(row=1, column=c).value]

            wb_diario = load_workbook(diario_path, data_only=True)
            ws_diario = wb_diario.active
            datas_diario = [ws_diario.cell(row=1, column=c).value for c in range(2, ws_diario.max_column + 1) if ws_diario.cell(row=1, column=c).value]

            str_cx2 = {d.strftime("%Y-%m-%d") for d in datas_cx2 if isinstance(d, datetime)}
            str_diario = {d.strftime("%Y-%m-%d") for d in datas_diario if isinstance(d, datetime)}

            self.pendentes = sorted(list(str_cx2 - str_diario))

            self.status["planilha_check"] = "success"
            return True

        except Exception as e:
            logger.error(f"Erro ao verificar paridade de planilhas: {str(e)}")
            self.status["planilha_check"] = f"error: {str(e)}"
            return False

    def get_status(self) -> dict:
        """Return the current status of the Cortex Padroeira"""
        return {
            "status": self.status,
            "dados": self.dados,
            "pendentes": self.pendentes,
            "has_pending_dates": len(self.pendentes) > 0
        }

# Telegram Bot Handlers
cortex = CortexPadroeira()

@bot.message_handler(func=lambda message: re.search(r'\bfechar\b', message.text.lower().strip()))
def handle_fechar(message):
    """Handler para mensagens naturais contendo 'fechar'"""
    bot.reply_to(message, "🤖 *Córtex Lab:* Processando dados brutos do Saurus...")
    dados = cortex.extrair_dados_saurus()

    if not dados:
        bot.reply_to(message, "❌ Erro: Arquivo 'fechamento_caixa.txt' não encontrado na pasta do laboratório.")
        return

    msg = (
        f"📊 *FATURAMENTO TOTAL DO DIA*\n\n"
        f"• DINHEIRO: R$ {dados['dinheiro']}\n"
        f"• CRÉDITO: R$ {dados['credito']}\n"
        f"• DÉBITO: R$ {dados['debito']}\n"
        f"• *TOTAL DO SISTEMA: R$ {dados['total']}*\n\n"
        f"⚖️ *Métricas Equivalentes (Automação):*\n"
        f"• Refeição Quilo: {dados['peso_buf']} KG\n"
        f"• Sobremesa Quilo: {dados['peso_sob']} KG\n"
        f"• Número de Clientes: {dados['clientes']}\n\n"
        f"✍️ Sandra, preencha o *Movto_cx2.xlsx* no PC **incluindo a SANGRIA na linha 42**.\n"
        f"Quando terminar e salvar, digite 'ok' aqui para consolidar!"
    )
    bot.send_message(message.chat.id, msg, parse_mode="Markdown")

@bot.message_handler(func=lambda message: re.search(r'\bok\b', message.text.lower().strip()))
def handle_ok(message):
    """Handler para mensagens naturais contendo 'ok'"""
    bot.reply_to(message, "⚡ *Córtex Lab:* Comparando planilhas e checando paridade...")

    if not cortex.verificar_paridade_planilhas():
        bot.reply_to(message, "❌ Erro ao verificar paridade de planilhas.")
        return

    if cortex.pendentes:
        bot.send_message(message.chat.id, f"🔄 *Datas pendentes detectadas no Caixa 2:* {len(cortex.pendentes)} dia(s).")
    else:
        bot.send_message(message.chat.id, "✅ Paridade total encontrada! Nenhuma data nova detectada.")

    # Dispara os motores unificados
    bot.send_message(message.chat.id, "🚀 *Iniciando motores de reconciliação...*")

    # Importa e executa os motores
    try:
        from engine_consolidacao import EngineConsolidacao
        from motor_balancete import MotorBalancete

        # Fase 1: Engine de Consolidação
        engine = EngineConsolidacao()
        result_engine = engine.executar_motor_unificado()

        if result_engine["status"] != "success":
            bot.send_message(message.chat.id, f"❌ Erro na Fase 1 (Engine): {result_engine.get('error', 'Erro desconhecido')}")
            return

        # Fase 2: Motor Balancete
        motor = MotorBalancete()
        result_motor = motor.injetar_balancete()

        if result_motor["status"] != "success":
            bot.send_message(message.chat.id, f"❌ Erro na Fase 2 (Motor): {result_motor.get('error', 'Erro desconhecido')}")
            return

        bot.send_message(message.chat.id, "✅ *Reconciliação concluída com sucesso!*")

    except Exception as e:
        logger.error(f"Erro ao executar motores: {str(e)}")
        bot.send_message(message.chat.id, f"❌ Erro ao executar motores: {str(e)}")

if __name__ == "__main__":
    logger.info("[*] CÓRTEX PADROEIRA: Ativo e escutando o Telegram...")
    logger.info(f"[*] Usando mock_box em: {PASTA_VENDAS_ORIGEM}")
    bot.infinity_polling()