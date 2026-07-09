#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Engine de Consolidação - FASE 2: Transposição Diário -> Balancete (Async Version)
Adaptado para Async Reconciliation Architecture V2
"""

import os
from datetime import datetime
from openpyxl import load_workbook
import logging
from typing import Dict, Any, Optional

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger("MotorBalanceteAsync")

RAIZ_PADROEIRA = "/home/teco/Nuvens/Box/Padroeira"
PASTA_RESTAURANTE_ANO = os.path.join(RAIZ_PADROEIRA, "Restaurante", "A2026")

# ==============================================================================
# DICIONÁRIO DE TRANSPOSIÇÃO (Matriz Cruzada)
# Chave: Coluna Destino no Balancete | Valor: Linha Origem no Movto Diário
# ==============================================================================
MAPA_DIARIO_PAD = {
    'B': 3,   # Peso Buf.
    'C': 4,   # Peso Sob.
    'E': 6,   # Dinheiro
    'F': 10,  # MasterCard (Redecard crédito)
    'G': 11,  # RedeShop (Redecard debito)
    'H': 12,  # Visa Cred (Cielo Visa Credito)
    'I': 13,  # Visa Deb (Cielo Visa Débito)
    'J': 7,   # Cheques
    'K': 24,  # Tickets (Linha de Total dos tickets)
    'L': 27,  # Cvs/E (Pix / conta corrente)
    'N': 32,  # Pend.Dia
    'O': 33,  # Pend.Pg.
    'P': 36,  # Serviço
    'Q': 37   # Mov/Dia (Real)
}

class MotorBalanceteAsync:
    """
    Async version of Motor Balancete for integration with Async Reconciliation V2
    """

    def __init__(self, aamm: str = None):
        self.aamm = aamm or datetime.now().strftime("%y%m")
        self.status = {
            "file_check": None,
            "data_extraction": None,
            "transposition": None,
            "save": None
        }
        self.stats = {
            "days_processed": 0,
            "lines_modified": 0,
            "new_days_added": 0
        }

    def normalizar_dia(self, valor_data) -> Optional[int]:
        """Extrai apenas o número do dia (inteiro) da data do Diário"""
        if isinstance(valor_data, datetime):
            return int(valor_data.day)
        if isinstance(valor_data, str):
            try:
                return int(datetime.strptime(valor_data.strip().split()[0], "%d/%m/%Y").day)
            except:
                try:
                    return int(datetime.strptime(valor_data.strip().split()[0], "%Y-%m-%d").day)
                except:
                    pass
        return None

    def injetar_balancete(self) -> Dict[str, Any]:
        """
        Execute the balance sheet injection process
        """
        try:
            diario_path = os.path.join(PASTA_RESTAURANTE_ANO, f"Movto_diario.{self.aamm}.xlsx")
            pad_path = os.path.join(PASTA_RESTAURANTE_ANO, f"Pad{self.aamm}.xlsx")

            # Check if files exist
            if not os.path.exists(diario_path) or not os.path.exists(pad_path):
                error_msg = "Arquivos da Fase 2 (Diário ou Pad) não encontrados no Box"
                logger.error(error_msg)
                self.status["file_check"] = f"error: {error_msg}"
                return {"status": "error", "error": error_msg}

            self.status["file_check"] = "success"
            logger.info("[*] Iniciando Fase 2: Transposição de Matriz (Diário -> Balancete Pad)...")

            # 1. Carrega o Diário para LEITURA dos valores finais calculados
            wb_diario = load_workbook(diario_path, data_only=True)
            ws_diario = wb_diario.active

            # Extrai a carga útil de todos os dias disponíveis no Diário
            carga_por_dia = {}
            for col in range(2, ws_diario.max_column + 1):
                v_data = ws_diario.cell(row=1, column=col).value
                if v_data:
                    dia_int = self.normalizar_dia(v_data)
                    if dia_int:
                        carga_por_dia[dia_int] = {}
                        # Coleta os valores conforme o mapa
                        for col_pad, linha_diario in MAPA_DIARIO_PAD.items():
                            valor = ws_diario.cell(row=linha_diario, column=col).value
                            # Substitui None por 0 para não quebrar fórmulas matemáticas no Balancete
                            carga_por_dia[dia_int][col_pad] = valor if valor is not None else 0.0

            if not carga_por_dia:
                error_msg = "Nenhum dia válido encontrado no Diário Mensal"
                logger.error(error_msg)
                self.status["data_extraction"] = f"error: {error_msg}"
                return {"status": "error", "error": error_msg}

            logger.info(f"[*] Carga extraída do Diário! Dias capturados: {sorted(list(carga_por_dia.keys()))}")
            self.status["data_extraction"] = "success"
            self.stats["days_processed"] = len(carga_por_dia)

            # 2. Carrega o Balancete (Pad) em modo ESCRITA na aba "Movimento Diario"
            wb_pad = load_workbook(pad_path, data_only=False)

            # Garante que estamos escrevendo na aba certa, independente de ser a ativa
            nome_aba = "Movimento Diario"
            if nome_aba in wb_pad.sheetnames:
                ws_pad = wb_pad[nome_aba]
            else:
                ws_pad = wb_pad.active
                logger.warning(f"Aba '{nome_aba}' não encontrada. Usando a aba ativa: {ws_pad.title}")

            # 3. Varredura e Sobreposição Agressiva no Balancete
            linhas_modificadas = 0
            novos_dias = 0
            for dia, valores in carga_por_dia.items():
                linha_destino = None

                # Procura o dia na Coluna A (limite de segurança até a linha 31 para não pegar totais)
                for r in range(2, 32):
                    celula_dia = ws_pad.cell(row=r, column=1).value
                    if celula_dia == dia or celula_dia == str(dia):
                        linha_destino = r
                        break

                # Se não achou o dia, procura a primeira linha vazia na Coluna A
                if not linha_destino:
                    for r in range(2, 32):
                        if not ws_pad.cell(row=r, column=1).value:
                            linha_destino = r
                            ws_pad.cell(row=linha_destino, column=1, value=dia)
                            logger.info(f"    [+] Novo dia ({dia}) inserido na linha {linha_destino} do Balancete.")
                            novos_dias += 1
                            break

                if linha_destino:
                    logger.debug(f"    -> [Sobreposição Ativa] Injetando dados do Dia {dia} na Linha {linha_destino}")
                    for letra_coluna, valor in valores.items():
                        # Pula a coluna Q se você não quer que ela seja sobrescrita
                        if letra_coluna == 'Q':
                            continue

                        ws_pad[f"{letra_coluna}{linha_destino}"].value = valor

                    # FORÇA A FÓRMULA NA COLUNA Q (Ajuste a fórmula abaixo conforme sua necessidade)
                    ws_pad[f'Q{linha_destino}'].value = f'=SUM(B{linha_destino}:P{linha_destino})'
                    linhas_modificadas += 1

            self.stats["lines_modified"] = linhas_modificadas
            self.stats["new_days_added"] = novos_dias
            self.status["transposition"] = "success"

            # 4. Salva o Balancete
            if linhas_modificadas > 0:
                try:
                    wb_pad.save(pad_path)
                    logger.info(f"\n[+] Fase 2 Concluída! {linhas_modificadas} dias sobrepostos com sucesso no Balancete (Pad{self.aamm}.xlsx).")
                    self.status["save"] = "success"
                    return {
                        "status": "success",
                        "stats": self.stats,
                        "details": self.status
                    }
                except Exception as e:
                    error_msg = f"Erro ao salvar o Balancete: {str(e)}"
                    logger.error(error_msg)
                    self.status["save"] = f"error: {str(e)}"
                    return {"status": "error", "error": error_msg}
            else:
                logger.info("Nenhuma modificação necessária no Balancete.")
                self.status["save"] = "skipped"
                return {
                    "status": "success",
                    "stats": self.stats,
                    "details": self.status
                }

        except Exception as e:
            error_msg = f"Erro inesperado no motor de balancete: {str(e)}"
            logger.error(error_msg)
            return {"status": "error", "error": error_msg}

    def get_status(self) -> Dict[str, Any]:
        """Return the current status of the motor"""
        return {
            "status": self.status,
            "stats": self.stats,
            "aamm": self.aamm
        }

if __name__ == "__main__":
    logger.info("[*] =================================================================")
    logger.info("[*] PHASE 2: MOTOR DE TRANSPOSIÇÃO E SOBREPOSIÇÃO AGRESSIVA - Async Version")
    logger.info("[*] =================================================================")

    motor = MotorBalanceteAsync()
    result = motor.injetar_balancete()

    if result["status"] == "success":
        logger.info("Motor de balancete executado com sucesso!")
        logger.info(f"Estatísticas: {motor.get_status()['stats']}")
    else:
        logger.error(f"Erro no motor de balancete: {result['error']}")