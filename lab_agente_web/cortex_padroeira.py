#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Córtex Principal - Automação Ecossistema Padroeira
Orquestrador Central: Telegram Bot + Playwright + Inteligência de Matrizes
"""

import os
import re
from datetime import datetime
import telebot
from openpyxl import load_workbook

# CONFIGURAÇÃO DE AMBIENTE
TOKEN_TELEGRAM = "8890227531:AAEWbuEyzA3G2VImmYICJBK8ZqCUsBbq9Ro"  # ← SUBSTITUA PELO SEU TOKEN REAL

# AJUSTE OS CAMINHOS PARA SUA ESTRUTURA ATUAL
PASTA_LAB = os.path.expanduser("~/work_out/automacao_padroeira_v1")
PASTA_VENDAS_ORIGEM = "/home/teco/Nuvens/Box/Padroeira/Padroeira vendas"

bot = telebot.TeleBot(TOKEN_TELEGRAM)

def extrair_dados_saurus():
    """Lê o fechamento_caixa.txt e calcula os totais e equivalentes de peso"""
    txt_path = os.path.join(PASTA_LAB, "fechamento_caixa.txt")
    if not os.path.exists(txt_path):
        return None

    with open(txt_path, "r", encoding="utf-8") as f:
        conteudo = f.read()

    # Busca valores financeiros usando Regex simples
    dinheiro = re.search(r"DINHEIRO \(\d+\):\s+([\d\.]+)", conteudo)
    credito = re.search(r"CRÉDITO \(\d+\):\s+([\d\.]+)", conteudo)
    debito = re.search(r"DÉBITO \(\d+\):\s+([\d\.]+)", conteudo)
    total = re.search(r"TOTAL \(\d+\):\s+([\d\.,]+)", conteudo)
    
    # Busca Pesos e Clientes
    clientes = re.search(r"Qtd\. Vendas\s+:\s+(\d+)", conteudo)
    peso_buf = re.search(r"REFEICAO QUILO\s+KG\s+([\d\.,]+)", conteudo)
    peso_sob = re.search(r"SOBREMESA QUILO\s+KG\s+([\d\.,]+)", conteudo)

    return {
        "dinheiro": dinheiro.group(1) if dinheiro else "0.00",
        "credito": credito.group(1) if credito else "0.00",
        "debito": debito.group(1) if debito else "0.00",
        "total": total.group(1).replace(",", ".") if total else "0.00",
        "clientes": clientes.group(1) if clientes else "0",
        "peso_buf": peso_buf.group(1).replace(",", ".") if peso_buf else "0.000",
        "peso_sob": peso_sob.group(1).replace(",", ".") if peso_sob else "0.000"
    }

@bot.message_handler(commands=['fechar'])
def comando_fechar(message):
    bot.reply_to(message, "🤖 *Córtex Lab:* Processando dados brutos do Saurus...")
    dados = extrair_dados_saurus()
    
    if not dados:
        bot.reply_to(message, "❌ Erro: Arquivo 'fechamento_caixa.txt' não encontrado na pasta do laboratório.")
        return

    msg = (
        f"📊 *FATURAMENTO TOTAL DO DIA*\n\n"
        f"• DINHEIRO: R$ {dados['dinheiro']}\n"
        f"• CRÉDITO: R$ {dados['credito']}\n"
        f"• DÉBITO: R$ {dados['debito']}\n"
        f"• *TOTAL DO SISTEMA: R$ {dados['total']}*\n\n"
        f"⚖️ *Métricas Equivalentes (Automação):*\n"
        f"• Refeição Quilo: {dados['peso_buf']} KG\n"
        f"• Sobremesa Quilo: {dados['peso_sob']} KG\n"
        f"• Número de Clientes: {dados['clientes']}\n\n"
        f"✍️ Sandra, preencha o *Movto_cx2.xlsx* no PC.\n"
        f"Quando terminar e salvar, digite */ok* aqui para consolidar!"
    )
    bot.send_message(message.chat.id, msg, parse_mode="Markdown")

@bot.message_handler(commands=['ok'])
def comando_ok(message):
    bot.reply_to(message, "⚡ *Córtex Lab:* Comparando planilhas e checando paridade...")
    
    cx2_path = os.path.join(PASTA_VENDAS_ORIGEM, "Movto_cx2.xlsx")
    aamm_atual = datetime.now().strftime("%y%m")
    diario_path = os.path.join(PASTA_LAB, f"Movto_diario.{aamm_atual}.xlsx")
    
    if not os.path.exists(diario_path):
        bot.reply_to(message, f"❌ Planilha de teste '{os.path.basename(diario_path)}' não encontrada no lab.")
        return

    wb_cx2 = load_workbook(cx2_path, data_only=True)
    ws_cx2 = wb_cx2.active
    datas_cx2 = [ws_cx2.cell(row=1, column=c).value for c in range(2, ws_cx2.max_column + 1) if ws_cx2.cell(row=1, column=c).value]

    wb_diario = load_workbook(diario_path, data_only=True)
    ws_diario = wb_diario.active
    datas_diario = [ws_diario.cell(row=1, column=c).value for c in range(2, ws_diario.max_column + 1) if ws_diario.cell(row=1, column=c).value]

    str_cx2 = {d.strftime("%Y-%m-%d") for d in datas_cx2 if isinstance(d, datetime)}
    str_diario = {d.strftime("%Y-%m-%d") for d in datas_diario if isinstance(d, datetime)}
    
    pendentes = sorted(list(str_cx2 - str_diario))

    if pendentes:
        bot.send_message(message.chat.id, f"🔄 *Datas pendentes detectadas no Caixa 2:* {len(pendentes)} dia(s). Rodando motores...")
    else:
        bot.send_message(message.chat.id, "✅ Paridade total encontrada! Nenhuma data nova detectada, mas os motores serão executados para garantir a consistência.")

    try:
        from engine_consolidacao import executar_motor_unificado
        from motor_balancete import injetar_balancete
        
        executar_motor_unificado()
        injetar_balancete()
        
        bot.send_message(message.chat.id, "🚀 *CONSOLIDAÇÃO DE TESTE CONCLUÍDA!* Verifique os arquivos no lab.")
    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Erro na execução dos scripts: {e}")

if __name__ == "__main__":
    print("[*] LAB CÓRTEX: Ativo e escutando o Telegram...")
    bot.infinity_polling()
