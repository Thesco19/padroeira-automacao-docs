#!/usr/bin/env python3
"""Teste da verificação de paridade entre as planilhas de mock."""

import os
from datetime import datetime
from openpyxl import load_workbook

PASTA_LAB = os.path.expanduser("~/work_out")
PASTA_VENDAS_ORIGEM = os.path.join(PASTA_LAB, "mock_box", "Padroeira_vendas")
PASTA_RESTAURANTE_ANO = os.path.join(PASTA_LAB, "mock_box", "Restaurante", "A2026")

def normalizar_data(valor):
    """Retorna um objeto date para casamento seguro de colunas"""
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, str):
        try:
            return datetime.strptime(valor.strip().split()[0], "%d/%m/%Y").date()
        except:
            try:
                return datetime.strptime(valor.strip().split()[0], "%Y-%m-%d").date()
            except:
                return None
    return None

def verificar_paridade_planilhas():
    """Verifica a paridade entre as planilhas Movto_cx2.xlsx e Movto_diario"""
    try:
        cx2_path = os.path.join(PASTA_VENDAS_ORIGEM, "Movto_cx2.xlsx")
        aamm_atual = datetime.now().strftime("%y%m")
        diario_path = os.path.join(PASTA_RESTAURANTE_ANO, f"Movto_diario.{aamm_atual}.xlsx")

        print(f"Caminho Caixa 2: {cx2_path}")
        print(f"Caminho Diário: {diario_path}")
        print(f"Existe Caixa 2: {os.path.exists(cx2_path)}")
        print(f"Existe Diário: {os.path.exists(diario_path)}")

        if not os.path.exists(diario_path):
            print(f"Planilha de teste '{os.path.basename(diario_path)}' não encontrada no mock_box")
            return False

        wb_cx2 = load_workbook(cx2_path, data_only=True)
        ws_cx2 = wb_cx2.active
        datas_cx2 = [ws_cx2.cell(row=1, column=c).value for c in range(2, ws_cx2.max_column + 1) if ws_cx2.cell(row=1, column=c).value]

        wb_diario = load_workbook(diario_path, data_only=True)
        ws_diario = wb_diario.active
        datas_diario = [ws_diario.cell(row=1, column=c).value for c in range(2, ws_diario.max_column + 1) if ws_diario.cell(row=1, column=c).value]

        print(f"Datas no Caixa 2: {datas_cx2}")
        print(f"Datas no Diário: {datas_diario}")

        str_cx2 = {d.strftime("%Y-%m-%d") for d in datas_cx2 if isinstance(d, datetime)}
        str_diario = {d.strftime("%Y-%m-%d") for d in datas_diario if isinstance(d, datetime)}

        print(f"Datas normalizadas Caixa 2: {str_cx2}")
        print(f"Datas normalizadas Diário: {str_diario}")

        pendentes = sorted(list(str_cx2 - str_diario))
        print(f"Datas pendentes: {pendentes}")

        return True

    except Exception as e:
        print(f"Erro ao verificar paridade de planilhas: {str(e)}")
        return False

if __name__ == "__main__":
    verificar_paridade_planilhas()