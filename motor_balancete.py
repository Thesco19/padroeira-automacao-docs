#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Motor de Balancete - FASE 2: Transposição Diário -> Balancete
Versão para Arquitetura de Reconciliação V2

Requisitos implementados:
1. Auditoria por Faturamento: Usa a linha 37 do Md (faturamento) como âncora de varredura,
   em vez de uma sequência cega de dias. Para cada coluna com faturamento ativo, localiza
   a linha correspondente no Ba (via Coluna A) e executa a sobreposição agressiva.
2. Mapeamento da Sangria: MAPA_DIARIO_PAD inclui 'R': 42 (Coluna R do Balancete <- Linha 42 do Diário).
3. Isolamento da Coluna Q: Nunca escreve valor fixo na coluna Q (Faturamento). Força a
   fórmula original =SUM(B{linha}:P{linha}) para manter o cálculo nativo.
"""

import os
from datetime import datetime
from openpyxl import load_workbook
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger("MotorBalanceteV2")

# AJUSTE PARA MOCK_BOX DURANTE DESENVOLVIMENTO
PASTA_LAB = os.path.expanduser("~/work_out")
PASTA_RESTAURANTE_ANO = os.path.join(PASTA_LAB, "mock_box", "Restaurante", "A2026")

# ==============================================================================
# DICIONÁRIO DE TRANSPOSIÇÃO (Matriz Cruzada)
# Chave: Coluna Destino no Balancete | Valor: Linha Origem no Movto Diário
# ==============================================================================
MAPA_DIARIO_PAD = {
    'B': 3,   # Peso Buf.
    'C': 4,   # Peso Sob.
    'E': 6,   # Dinheiro
    'F': 10,  # MasterCard (Redecard crédito)
    'G': 11,  # RedeShop (Redecard débito)
    'H': 12,  # Visa Cred (Cielo Visa Crédito)
    'I': 13,  # Visa Deb (Cielo Visa Débito)
    'J': 7,   # Cheques
    'K': 24,  # Tickets (Linha de Total dos tickets)
    'L': 27,  # Cvs/E (Pix / conta corrente)
    'N': 32,  # Pend.Dia
    'O': 33,  # Pend.Pg.
    'P': 36,  # Serviço
    'Q': 37,  # Mov/Dia (Real) — Faturamento (âncora)
    'R': 42   # Sangria (Linha 42 do Diário -> Coluna R do Balancete)
}

# Linha de faturamento no Diário (âncora de varredura)
LINHA_FATURAMENTO_MD = 37


class MotorBalancete:
    """
    Motor Balancete para a Arquitetura de Reconciliação V2
    """

    def __init__(self, aamm: str = None):
        self.aamm = aamm or datetime.now().strftime("%y%m")
        self.status = {
            "file_check": None,
            "data_extraction": None,
            "transposition": None,
            "save": None
        }
        self.stats = {
            "days_processed": 0,
            "lines_modified": 0,
            "new_days_added": 0
        }

    def normalizar_dia(self, valor_data) -> int:
        """Extrai apenas o número do dia (inteiro) da data do Diário"""
        if isinstance(valor_data, datetime):
            return int(valor_data.day)
        if isinstance(valor_data, str):
            try:
                return int(datetime.strptime(valor_data.strip().split()[0], "%d/%m/%Y").day)
            except:
                try:
                    return int(datetime.strptime(valor_data.strip().split()[0], "%Y-%m-%d").day)
                except:
                    pass
        return None

    def injetar_balancete(self) -> dict:
        """
        Executa a transposição Diário -> Balancete usando auditoria por faturamento.

        Returns:
            dict: {"status": "success"|"error", "error": str, "stats": dict}
        """
        try:
            diario_path = os.path.join(PASTA_RESTAURANTE_ANO, f"Movto_diario.{self.aamm}.xlsx")
            pad_path = os.path.join(PASTA_RESTAURANTE_ANO, f"Pad{self.aamm}.xlsx")

            # Check if files exist
            if not os.path.exists(diario_path) or not os.path.exists(pad_path):
                error_msg = "Arquivos da Fase 2 (Diário ou Pad) não encontrados no mock_box"
                logger.error(error_msg)
                self.status["file_check"] = f"error: {error_msg}"
                return {"status": "error", "error": error_msg}

            self.status["file_check"] = "success"
            logger.info("[*] Iniciando Fase 2: Transposição de Matriz (Diário -> Balancete Pad)...")
            logger.info(f"[*] Ancoragem por faturamento (linha {LINHA_FATURAMENTO_MD} do Diário)")

            # 1. Carrega o Diário para LEITURA dos valores finais calculados
            wb_diario = load_workbook(diario_path, data_only=True)
            ws_diario = wb_diario.active

            # ==================================================================
            # AUDITORIA POR FATURAMENTO (linha 37 como âncora)
            # Em vez de sequência cega de dias, só processa colunas com faturamento ativo
            # ==================================================================
            carga_por_dia = {}
            dias_ancorados = []

            for col in range(2, ws_diario.max_column + 1):
                v_data = ws_diario.cell(row=1, column=col).value
                if not v_data:
                    continue

                dia_int = self.normalizar_dia(v_data)
                if not dia_int:
                    continue

                # ÂNCORA: verifica o faturamento na linha 37 do Diário
                faturamento = ws_diario.cell(row=LINHA_FATURAMENTO_MD, column=col).value

                # Só processa colunas com faturamento ATIVO (não None e não zero)
                if faturamento is None:
                    continue
                try:
                    if float(faturamento) == 0:
                        continue
                except (TypeError, ValueError):
                    continue

                # Faturamento ativo: coleta a carga útil conforme o mapa
                dias_ancorados.append(dia_int)
                carga_por_dia[dia_int] = {}
                for col_pad, linha_diario in MAPA_DIARIO_PAD.items():
                    valor = ws_diario.cell(row=linha_diario, column=col).value
                    # Substitui None por 0 para não quebrar fórmulas no Balancete
                    carga_por_dia[dia_int][col_pad] = valor if valor is not None else 0.0

            if not carga_por_dia:
                error_msg = "Nenhum dia com faturamento ativo encontrado no Diário Mensal"
                logger.error(error_msg)
                self.status["data_extraction"] = f"error: {error_msg}"
                return {"status": "error", "error": error_msg}

            logger.info(f"[*] Carga extraída via âncora de faturamento! Dias capturados: {sorted(dias_ancorados)}")
            self.status["data_extraction"] = "success"
            self.stats["days_processed"] = len(carga_por_dia)

            # 2. Carrega o Balancete (Pad) em modo ESCRITA na aba "Movimento Diario"
            wb_pad = load_workbook(pad_path, data_only=False)

            # Garante que estamos escrevendo na aba certa, independente de ser a ativa
            nome_aba = "Movimento Diario"
            if nome_aba in wb_pad.sheetnames:
                ws_pad = wb_pad[nome_aba]
            else:
                ws_pad = wb_pad.active
                logger.warning(f"Aba '{nome_aba}' não encontrada. Usando a aba ativa: {ws_pad.title}")

            # ==================================================================
            # 3. VARREDURA E SOBREPOSIÇÃO AGRESSIVA NO BALANCETE
            # ==================================================================
            linhas_modificadas = 0
            novos_dias = 0

            for dia, valores in carga_por_dia.items():
                linha_destino = None

                # Procura o dia na Coluna A (limite de segurança até a linha 31 p/ não pegar totais)
                for r in range(2, 32):
                    celula_dia = ws_pad.cell(row=r, column=1).value
                    # Tenta casar como int, str ou extraído de data
                    if celula_dia == dia or celula_dia == str(dia):
                        linha_destino = r
                        break
                    # Caso a Coluna A tenha uma data completa, extrai o dia
                    dia_extraido = self.normalizar_dia(celula_dia)
                    if dia_extraido == dia:
                        linha_destino = r
                        break

                # Se não achou o dia, procura a primeira linha vazia na Coluna A
                if not linha_destino:
                    for r in range(2, 32):
                        if not ws_pad.cell(row=r, column=1).value:
                            linha_destino = r
                            ws_pad.cell(row=linha_destino, column=1, value=dia)
                            logger.info(f"    [+] Novo dia ({dia}) inserido na linha {linha_destino} do Balancete.")
                            novos_dias += 1
                            break

                if linha_destino:
                    logger.debug(f"    -> [Sobreposição Ativa] Injetando dados do Dia {dia} na Linha {linha_destino}")
                    for letra_coluna, valor in valores.items():
                        # ISOLAMENTO DA COLUNA Q: nunca escreve valor fixo no Faturamento
                        # A fórmula =SUM(B: P) é forçada abaixo para manter o cálculo nativo
                        if letra_coluna == 'Q':
                            continue

                        ws_pad[f"{letra_coluna}{linha_destino}"].value = valor

                    # FORÇA A FÓRMULA NA COLUNA Q (Faturamento) — cálculo nativo
                    ws_pad[f'Q{linha_destino}'].value = f'=SUM(B{linha_destino}:P{linha_destino})'
                    linhas_modificadas += 1

            self.stats["lines_modified"] = linhas_modificadas
            self.stats["new_days_added"] = novos_dias
            self.status["transposition"] = "success"

            # 4. Salva o Balancete
            if linhas_modificadas > 0:
                try:
                    wb_pad.save(pad_path)
                    logger.info(f"\n[+] Fase 2 Concluída! {linhas_modificadas} dias sobrepostos com sucesso no Balancete (Pad{self.aamm}.xlsx).")
                    self.status["save"] = "success"
                    return {
                        "status": "success",
                        "stats": self.stats,
                        "details": self.status
                    }
                except Exception as e:
                    error_msg = f"Erro ao salvar o Balancete: {str(e)}"
                    logger.error(error_msg)
                    self.status["save"] = f"error: {str(e)}"
                    return {"status": "error", "error": error_msg}
            else:
                logger.info("Nenhuma modificação necessária no Balancete.")
                self.status["save"] = "skipped"
                return {
                    "status": "success",
                    "stats": self.stats,
                    "details": self.status
                }

        except Exception as e:
            error_msg = f"Erro inesperado no motor de balancete: {str(e)}"
            logger.error(error_msg)
            return {"status": "error", "error": error_msg}

    def get_status(self) -> dict:
        """Return the current status of the motor"""
        return {
            "status": self.status,
            "stats": self.stats,
            "aamm": self.aamm
        }

if __name__ == "__main__":
    logger.info("[*] =================================================================")
    logger.info("[*] PHASE 2: MOTOR DE TRANSPOSIÇÃO E SOBREPOSIÇÃO AGRESSIVA - V2")
    logger.info("[*] =================================================================")
    logger.info(f"[*] Usando mock_box em: {PASTA_RESTAURANTE_ANO}")

    motor = MotorBalancete()
    result = motor.injetar_balancete()

    if result["status"] == "success":
        logger.info("Motor de balancete executado com sucesso!")
        logger.info(f"Estatísticas: {motor.get_status()['stats']}")
    else:
        logger.error(f"Erro no motor de balancete: {result['error']}")